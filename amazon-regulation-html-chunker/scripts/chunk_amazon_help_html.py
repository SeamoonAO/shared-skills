from __future__ import annotations

import argparse
import copy
import json
import re
from datetime import datetime, timezone
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from bs4 import BeautifulSoup, FeatureNotFound, NavigableString, Tag


NOISE_TAGS = {"script", "style", "noscript", "svg", "iframe", "link", "meta"}
CONTENT_ROOT_SELECTORS = (
    ("id", "help-content"),
    ("class", "help-content"),
    ("id", "full-help-page"),
    ("class", "hh-help-page"),
)
NOISE_IDS = {
    "hh-sidebar",
    "hh-nav-tree-widget",
    "sc-footer-container",
    "skip-link",
}
NOISE_CLASS_PATTERNS = (
    "sidebar",
    "nav",
    "footer",
    "cookie",
    "search",
)
TITLE_EXCLUSIONS = {
    "本页面内容",
    "Page contents",
    "目录",
}
LAST_UPDATED_PATTERNS = (
    re.compile(r"最后更新时间[:：]\s*(.+)$"),
    re.compile(r"Last updated[:：]?\s*(.+)$", re.IGNORECASE),
    re.compile(r"Letzte Aktualisierung[:：]?\s*(.+)$", re.IGNORECASE),
)
TAIL_UPDATED_PATTERNS = (
    re.compile(r"^(?:\u6700\u540e\u66f4\u65b0(?:\u65f6\u95f4)?|\u4e0a\u6b21\u66f4\u65b0(?:\u65e5\u671f)?)\s*[:\uff1a]"),
    re.compile(r"^last updated\s*[:\uff1a]?", re.IGNORECASE),
    re.compile(r"^letzte aktualisierung\s*[:\uff1a]?", re.IGNORECASE),
)
TAIL_DISCLAIMER_PATTERNS = (
    re.compile(r"^\u514d\u8d23\u58f0\u660e\s*[:\uff1a]"),
    re.compile(r"\u672c\u9875\u9762\u63d0\u4f9b\u7684\u4fe1\u606f\u4e0d\u6784\u6210"),
    re.compile(r"\u65e0\u610f\u6784\u6210\u6cd5\u5f8b\u5efa\u8bae"),
    re.compile(r"\u4ec5\u4f9b\u4e00\u822c\u53c2\u8003"),
    re.compile(r"\u5efa\u8bae\u54a8\u8be2\u6cd5\u5f8b\u987e\u95ee"),
    re.compile(r"^disclaimer\s*[:\uff1a]?", re.IGNORECASE),
    re.compile(r"does not constitute", re.IGNORECASE),
    re.compile(r"legal advice", re.IGNORECASE),
    re.compile(r"for general informational purposes only", re.IGNORECASE),
    re.compile(r"consult legal counsel", re.IGNORECASE),
    re.compile(r"^haftungsausschluss\s*[:\uff1a]?", re.IGNORECASE),
    re.compile(r"keine rechtsberatung", re.IGNORECASE),
    re.compile(r"allgemeine information", re.IGNORECASE),
    re.compile(r"rechtsberater", re.IGNORECASE),
)
LEADING_NOISE_PATTERNS = (
    re.compile(
        r"^\u4ee5\u4e0b\u5185\u5bb9\u4ec5\u4f5c\u4e3a\u6307\u5357\u3002\u4e9a\u9a6c\u900a\u4fdd\u7559\u4ee5\u4efb\u4f55\u7406\u7531\u79fb\u9664\u5176\u8ba4\u4e3a\u4e0d\u9002\u5408\u9500\u552e\u7684\u4efb\u4f55\u5546\u54c1(?:\u4fe1\u606f)?\u7684\u6743\u5229\u3002?$"
    ),
    re.compile(
        r"^\u60a8\u5c1d\u8bd5\u67e5\u770b\u7684\u9875\u9762\u5c1a\u672a\u63d0\u4f9b\u6240\u9009\u8bed\u8a00\u7248\u672c\u3002\u8bf7\u67e5\u770b\u6b64\u9875\u9762\u7684\u5176\u4ed6\u8bed\u8a00\u7248\u672c\u3002?$"
    ),
)
TOKEN_PATTERN = re.compile(r"[A-Za-z][A-Za-z0-9/_-]{2,}|[\u4e00-\u9fff]{2,}")
FRONTMATTER_PATTERN = re.compile(r"^---\s*\n(.*?)\n---\s*\n(.*)$", re.DOTALL)
DEFAULT_STOPWORDS = {
    "the",
    "and",
    "for",
    "with",
    "from",
    "that",
    "this",
    "into",
    "your",
    "you",
    "are",
    "not",
    "can",
    "all",
    "may",
    "will",
    "must",
    "should",
    "amazon",
    "seller",
    "central",
    "help",
    "policy",
    "policies",
    "page",
    "pages",
    "use",
    "used",
    "using",
    "please",
    "商品",
    "产品",
    "页面",
    "相关",
    "提供",
    "信息",
}
RISK_TERM_RULES = (
    (re.compile(r"(weapon|firearm|rifle|gun|ammunition|武器|枪|火器)", re.IGNORECASE), "weapon-related product"),
    (re.compile(r"(battery|lithium|锂电|电池)", re.IGNORECASE), "battery safety risk"),
    (re.compile(r"(chemical|toxic|hazardous|化学|有毒|危险品)", re.IGNORECASE), "chemical safety risk"),
    (re.compile(r"(medical|drug|medicine|疗效|医疗|药品)", re.IGNORECASE), "medical/drug compliance risk"),
    (re.compile(r"(adult|sexual|性爱|成人用品)", re.IGNORECASE), "adult product risk"),
    (re.compile(r"(food|beverage|edible|食品|饮料|可食用)", re.IGNORECASE), "food ingestion risk"),
)
SCENARIO_TERM_RULES = (
    (re.compile(r"(child|children|kid|toy|婴儿|儿童|玩具)", re.IGNORECASE), "children use scenario"),
    (re.compile(r"(pet|animal|宠物|动物)", re.IGNORECASE), "pet use scenario"),
    (re.compile(r"(household|home|kitchen|家用|厨房)", re.IGNORECASE), "household use scenario"),
    (re.compile(r"(outdoor|camping|hunting|户外|露营)", re.IGNORECASE), "outdoor use scenario"),
    (re.compile(r"(body|skin|oral|topical|皮肤|口服)", re.IGNORECASE), "body-contact scenario"),
)
SEMANTIC_VARIANTS = {
    "weapon-related product": ["firearm accessory", "restricted weapon component"],
    "battery safety risk": ["lithium battery compliance", "battery hazard policy"],
    "chemical safety risk": ["hazmat restriction", "chemical exposure risk"],
    "medical/drug compliance risk": ["medical claim policy", "drug-related restriction"],
    "adult product risk": ["sexual wellness policy", "adult listing restriction"],
    "food ingestion risk": ["ingestible product policy", "food-contact compliance"],
    "children use scenario": ["child safety policy", "toy compliance"],
}
REGULATION_BUCKET_RULES = (
    (
        "regulatory_anchors",
        (
            re.compile(r"(article|section|annex|directive|regulation|policy|法案|法规|政策|条款|章节|附录)", re.IGNORECASE),
            re.compile(r"(G[A-Z0-9]{6,}|art\.?\s*\d+|section\s+\d+)", re.IGNORECASE),
        ),
    ),
    (
        "scope_and_actors",
        (
            re.compile(r"(seller|buyer|manufacturer|distributor|retailer|importer|exporter|销售伙伴|制造商|分销商|零售商|进口商|出口商)", re.IGNORECASE),
            re.compile(r"(marketplace|jurisdiction|state|country|region|州|国家|地区|市场)", re.IGNORECASE),
        ),
    ),
    (
        "obligations_and_prohibitions",
        (
            re.compile(r"(must|must not|may not|required|prohibited|forbidden|不得|必须|禁止|不允许|应当)", re.IGNORECASE),
        ),
    ),
    (
        "conditions_and_thresholds",
        (
            re.compile(r"(if|when|unless|before|after|threshold|limit|frequency|date|deadline|除非|如果|当|之前|之后|阈值|限制|日期|截止)", re.IGNORECASE),
            re.compile(r"(\d+\s*(mg|kg|g|ml|l|%|days?|months?|years?|天|月|年))", re.IGNORECASE),
        ),
    ),
    (
        "exceptions_and_exemptions",
        (
            re.compile(r"(except|exception|exempt|safe harbor|豁免|例外|除外|免责)", re.IGNORECASE),
        ),
    ),
    (
        "evidence_and_documents",
        (
            re.compile(r"(certificate|report|declaration|form|record|document|permit|证书|报告|声明|表格|记录|文件|许可证)", re.IGNORECASE),
        ),
    ),
    (
        "enforcement_and_consequences",
        (
            re.compile(r"(penalty|fine|suspend|suspension|remove|removal|violation|违规|罚款|处罚|暂停|移除)", re.IGNORECASE),
        ),
    ),
    (
        "domain_entities",
        (
            re.compile(r"(battery|chemical|drug|food|beverage|weapon|animal|toy|medical|lithium|电池|化学|药品|食品|饮料|武器|动物|玩具|医疗)", re.IGNORECASE),
        ),
    ),
)


def default_output_root() -> Path:
    # scripts/chunk_amazon_help_html.py -> amazon-regulation-html-chunker/output
    return Path(__file__).resolve().parents[1] / "output"


def default_index_output_path(input_path: Path, output_root: Path) -> Path:
    token = input_path.stem if input_path.is_file() else (input_path.name or "batch")
    return output_root / f"{token}.index-seed.json"


def default_llm_packets_output_path(input_path: Path, output_root: Path) -> Path:
    token = input_path.stem if input_path.is_file() else (input_path.name or "batch")
    return output_root / f"{token}.llm-doc-packets.jsonl"


def default_llm_merged_output_path(results_input: Path, output_root: Path) -> Path:
    return output_root / f"{results_input.stem}.llm-merged.json"


def default_excel_output_path(merged_input: Path, output_root: Path) -> Path:
    return output_root / f"{merged_input.stem}.xlsx"


def default_combined_review_output_path(semantic_input: Path, output_root: Path) -> Path:
    return output_root / f"{semantic_input.stem}.combined-review.xlsx"


def default_semantic_merged_output_path(results_input: Path, output_root: Path) -> Path:
    return output_root / f"{results_input.stem}.semantic-merged.json"


def default_semantic_excel_output_path(merged_input: Path, output_root: Path) -> Path:
    return output_root / f"{merged_input.stem}.semantic.xlsx"


def default_index_packets_output_path(semantic_input: Path, output_root: Path) -> Path:
    return output_root / f"{semantic_input.stem}.index-chunk-packets.jsonl"


def default_manifest_output_path(input_path: Path, output_root: Path) -> Path:
    return output_root / f"{input_path.stem}.task-manifest.json"


def artifact_family_stem(input_path: Path) -> str:
    stem = input_path.stem
    for suffix in (".llm-doc-packets", ".doc-packets", ".index-chunk-packets", ".index-packets"):
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def default_prompt_preview_dir(input_path: Path, output_root: Path) -> Path:
    return output_root / f"{artifact_family_stem(input_path)}.prompt-preview"


def default_runner_input_dir(input_path: Path, output_root: Path) -> Path:
    return output_root / f"{artifact_family_stem(input_path)}.runner-inputs"


def stage_a_prompt_template_path() -> str:
    return str(Path(__file__).resolve().parents[1] / "references" / "subagent-stage-a-semantic-chunk-prompt.md")


def stage_b_prompt_template_path() -> str:
    return str(Path(__file__).resolve().parents[1] / "references" / "subagent-stage-b-inverted-index-prompt.md")


SUBAGENT_MODEL_ALIASES = {
    "gpt-5.3-codex-spark": "gpt-5.3-codex-spark",
    "gpt-5.3 codex spark": "gpt-5.3-codex-spark",
    "gpt-5.3-codex": "gpt-5.3-codex",
    "gpt-5.3 codex": "gpt-5.3-codex",
    "gpt-5.4-mini": "gpt-5.4-mini",
    "gpt-5.4 mini": "gpt-5.4-mini",
    "gpt-5.4": "gpt-5.4",
    "gpt-5.2": "gpt-5.2",
}
DEFAULT_MIN_INDEX_SCORE = 4
NO_FALLBACK_TOKENS = {"", "none", "null", "false", "off"}


def default_subagent_model() -> str:
    return "gpt-5.3-codex-spark"


def default_fallback_subagent_models(primary_model: str | None = None) -> list[str]:
    resolved_primary = normalize_subagent_model(primary_model or default_subagent_model())
    if resolved_primary == "gpt-5.3-codex-spark":
        return ["gpt-5.4-mini"]
    return []


def normalize_subagent_model(value: Any) -> str:
    text = normalize_text(str(value or ""))
    if not text:
        return default_subagent_model()
    key = text.casefold()
    if key in SUBAGENT_MODEL_ALIASES:
        return SUBAGENT_MODEL_ALIASES[key]
    raise SystemExit(
        "Unsupported --subagent-model. Supported values include: "
        "GPT-5.3-Codex-Spark, GPT-5.4-Mini, GPT-5.4, GPT-5.3-Codex, GPT-5.2."
    )


def normalize_subagent_model_list(value: Any, primary_model: str | None = None) -> list[str]:
    resolved_primary = normalize_subagent_model(primary_model or default_subagent_model())
    if value is None:
        candidates = default_fallback_subagent_models(resolved_primary)
    elif isinstance(value, list):
        candidates = [str(item) for item in value]
    else:
        text = normalize_text(str(value))
        if text.casefold() in NO_FALLBACK_TOKENS:
            candidates = []
        else:
            candidates = re.split(r"[,\n|]+", text)

    normalized: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        item = normalize_text(str(candidate))
        if not item:
            continue
        resolved = normalize_subagent_model(item)
        if resolved == resolved_primary or resolved in seen:
            continue
        seen.add(resolved)
        normalized.append(resolved)
    return normalized


def normalize_min_index_score(value: Any) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        raise SystemExit("--min-index-score must be an integer between 0 and 10.")
    if not 0 <= parsed <= 10:
        raise SystemExit("--min-index-score must be an integer between 0 and 10.")
    return parsed


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_semantic_text(value: str) -> str:
    return normalize_text(value).casefold()


def unique_preserving_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        candidate = normalize_text(value)
        if not candidate:
            continue
        key = candidate.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(candidate)
    return result


def normalize_term_list(values: Any) -> list[str]:
    if not isinstance(values, list):
        return []
    normalized = [normalize_text(str(value)) for value in values if normalize_text(str(value))]
    return unique_preserving_order(normalized)


def tokenize_text(text: str) -> list[str]:
    return [match.group(0) for match in TOKEN_PATTERN.finditer(text or "")]


def build_term_candidates(text: str, limit: int = 10) -> list[str]:
    counts: Counter[str] = Counter()
    for token in tokenize_text(text):
        normalized = token.casefold()
        if normalized in DEFAULT_STOPWORDS:
            continue
        if len(normalized) <= 2:
            continue
        counts[normalized] += 1

    ranked = [token for token, _ in counts.most_common(limit)]
    return unique_preserving_order(ranked)


def infer_regulation_bucket_hints(*parts: str) -> list[str]:
    text = " ".join(normalize_text(part) for part in parts if normalize_text(part))
    if not text:
        return []

    hints: list[str] = []
    for bucket_name, patterns in REGULATION_BUCKET_RULES:
        if any(pattern.search(text) for pattern in patterns):
            hints.append(bucket_name)
    return hints


def normalize_screening_relevance_score(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return max(0, min(10, int(value)))
    except (TypeError, ValueError):
        return None


def extract_global_entities(page: dict, limit: int = 15) -> list[str]:
    source_parts = [page.get("title", "")]
    for block in page.get("blocks", []):
        text = normalize_text(block.get("text", ""))
        if text:
            source_parts.append(text)
    return build_term_candidates(" ".join(source_parts), limit=limit)


def build_page_summary(page: dict, max_blocks: int = 3, max_chars: int = 360) -> str:
    snippets: list[str] = []
    for block in page.get("blocks", []):
        if block.get("type") == "heading":
            continue
        text = normalize_text(block.get("text", ""))
        if not text:
            continue
        snippets.append(text)
        if len(snippets) >= max_blocks:
            break

    summary = " ".join(snippets)
    summary = normalize_text(summary)
    if len(summary) > max_chars:
        return summary[: max_chars - 3].rstrip() + "..."
    return summary


def parse_frontmatter_markdown(markdown_text: str) -> tuple[dict[str, str], str]:
    normalized = markdown_text.replace("\r\n", "\n")
    match = FRONTMATTER_PATTERN.match(normalized)
    if not match:
        return {}, normalized

    frontmatter, body = match.groups()
    metadata: dict[str, str] = {}
    for line in frontmatter.splitlines():
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        metadata[key.strip()] = value.strip()
    return metadata, body


def parse_markdown_heading(text: str) -> tuple[int, str] | None:
    stripped = text.lstrip()
    if not stripped.startswith("#"):
        return None

    prefix = stripped.split(" ", 1)[0]
    level = len(prefix)
    if level < 1 or level > 6:
        return None

    heading_text = stripped[level:].strip()
    if not heading_text:
        return None
    return level, heading_text


def parse_markdown_list_item(text: str) -> tuple[int, str] | None:
    if not text.strip().startswith("- "):
        return None

    leading_spaces = len(text) - len(text.lstrip(" "))
    indent = max(0, leading_spaces // 2)
    value = normalize_text(text.strip()[2:])
    if not value:
        return None
    return indent, value


def parse_markdown_draft(path: Path) -> dict:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    metadata, body = parse_frontmatter_markdown(raw)
    source_file = metadata.get("source_file", f"{path.stem}.html")
    title = metadata.get("title", Path(source_file).stem)
    marketplace = metadata.get("marketplace") or extract_marketplace(source_file)

    blocks: list[dict] = []
    heading_stack: list[str] = [title]

    for line in body.splitlines():
        if not line.strip():
            continue

        heading = parse_markdown_heading(line)
        if heading:
            level, heading_text = heading
            if level == 1:
                # Skip top-level page title heading.
                continue
            markdown_level = heading_level_to_markdown(level)
            relative_depth = max(0, markdown_level - 2)
            heading_stack[:] = heading_stack[: relative_depth + 1]
            heading_stack.append(heading_text)
            blocks.append(
                {
                    "type": "heading",
                    "level": markdown_level,
                    "text": heading_text,
                    "heading_path": " > ".join(heading_stack),
                    "anchor": None,
                }
            )
            continue

        list_item = parse_markdown_list_item(line)
        if list_item:
            indent, value = list_item
            blocks.append(
                {
                    "type": "list_item",
                    "text": value,
                    "indent": indent,
                    "heading_path": " > ".join(heading_stack),
                    "anchor": None,
                }
            )
            continue

        paragraph = normalize_text(line)
        if paragraph:
            blocks.append(
                {
                    "type": "paragraph",
                    "text": paragraph,
                    "heading_path": " > ".join(heading_stack),
                    "anchor": None,
                }
            )

    return {
        "source_file": source_file,
        "source_url": metadata.get("source_url"),
        "marketplace": marketplace,
        "title": title,
        "last_updated": None,
        "blocks": blocks,
    }


def read_html(path: str | Path) -> str:
    return Path(path).read_text(encoding="utf-8", errors="ignore")


def parse_html(html: str) -> BeautifulSoup:
    try:
        soup = BeautifulSoup(html, "lxml")
    except FeatureNotFound:
        soup = BeautifulSoup(html, "html.parser")
    for tag in soup.find_all(NOISE_TAGS):
        tag.decompose()

    for element in soup.find_all(True):
        if should_drop_node(element):
            element.decompose()

    return soup


def should_drop_node(node: Tag) -> bool:
    if getattr(node, "attrs", None) is None:
        return False

    node_id = node.get("id") or ""
    if node_id in NOISE_IDS:
        return True

    joined_classes = " ".join(node.get("class") or []).lower()
    for pattern in NOISE_CLASS_PATTERNS:
        if pattern in joined_classes:
            return True

    role = (node.get("role") or "").lower()
    if role == "navigation":
        return True

    return False


def extract_content_root(html: str | BeautifulSoup) -> Tag:
    soup = html if isinstance(html, BeautifulSoup) else parse_html(html)

    for selector_type, selector_value in CONTENT_ROOT_SELECTORS:
        if selector_type == "id":
            node = soup.find(id=selector_value)
        else:
            node = soup.find(class_=selector_value)
        if node:
            return node

    return score_content_candidates(soup)


def score_content_candidates(soup: BeautifulSoup) -> Tag:
    best_node = soup.body or soup
    best_score = float("-inf")

    for node in soup.find_all(["main", "article", "section", "div"]):
        text = normalize_text(node.get_text(" ", strip=True))
        if len(text) < 200:
            continue

        headings = len(node.find_all(["h1", "h2", "h3", "h4"]))
        paragraphs = len(node.find_all(["p", "li"]))
        links = len(node.find_all("a"))
        score = len(text) + paragraphs * 80 + headings * 120 - links * 20

        identity = f"{node.get('id', '')} {' '.join(node.get('class') or [])}".lower()
        if any(token in identity for token in ("content", "article", "main", "help")):
            score += 300
        if any(token in identity for token in ("sidebar", "nav", "footer", "menu")):
            score -= 500

        if score > best_score:
            best_score = score
            best_node = node

    return best_node


def find_page_container(root: Tag) -> Tag:
    container = root.find_parent(id="full-help-page")
    if container:
        return container

    container = root.find_parent(class_="hh-help-page")
    if container:
        return container

    return root


def find_page_title(root: Tag, source_file: str) -> str:
    container = find_page_container(root)
    for heading in container.find_all(["h1", "h2", "h3"]):
        text = normalize_text(heading.get_text(" ", strip=True))
        if text and text not in TITLE_EXCLUSIONS:
            return text

    stem = Path(source_file).stem
    if "_" in stem:
        parts = stem.split("_")
        if len(parts) >= 2:
            return parts[1]
    return stem


def find_last_updated(container: Tag) -> str | None:
    for text_node in container.find_all(string=True):
        text = normalize_text(str(text_node))
        for pattern in LAST_UPDATED_PATTERNS:
            match = pattern.search(text)
            if match:
                return normalize_text(match.group(1))
    return None


def extract_marketplace(source_file: str) -> str | None:
    prefix = Path(source_file).stem.split("_", 1)[0]
    if prefix:
        return prefix
    return None


def extract_page_id(source_name: str) -> str | None:
    match = re.search(r"(G[A-Z0-9]+)", source_name)
    if match:
        return match.group(1)
    return None


def parse_index_file(index_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    filename_to_url: dict[str, str] = {}
    page_id_to_url: dict[str, str] = {}

    text = index_path.read_text(encoding="utf-8", errors="ignore")
    for line in text.splitlines():
        parts = [part.strip() for part in line.split("\t") if part.strip()]
        if len(parts) < 2:
            continue

        source_file = None
        source_url = None
        for part in parts:
            if part.endswith(".html"):
                source_file = part
            elif part.startswith("http://") or part.startswith("https://"):
                source_url = part

        if not source_file or not source_url:
            continue

        filename_to_url[source_file] = source_url
        page_id = extract_page_id(source_file)
        if page_id:
            page_id_to_url[page_id] = source_url

    return filename_to_url, page_id_to_url


def resolve_source_url(source_file: str, index_paths: Iterable[Path]) -> str | None:
    page_id = extract_page_id(source_file)

    for index_path in index_paths:
        filename_to_url, page_id_to_url = parse_index_file(index_path)
        if source_file in filename_to_url:
            return filename_to_url[source_file]
        if page_id and page_id in page_id_to_url:
            return page_id_to_url[page_id]

    return None


def discover_index_files_for_path(path: Path) -> list[Path]:
    search_dir = path.parent if path.is_file() else path
    candidates: list[Path] = []
    seen: set[Path] = set()

    prefix = extract_marketplace(path.name if path.is_file() else path.name)

    def add_if_exists(candidate: Path) -> None:
        if candidate.exists() and candidate.is_file():
            resolved = candidate.resolve()
            if resolved not in seen:
                seen.add(resolved)
                candidates.append(resolved)

    for candidate in search_dir.glob("*_index.txt"):
        add_if_exists(candidate)

    if prefix:
        for parent in [search_dir, *search_dir.parents]:
            add_if_exists(parent / f"{prefix}_index.txt")

    return candidates


def resolve_source_url_for_file(path: Path) -> str | None:
    return resolve_source_url(path.name, discover_index_files_for_path(path))


def heading_level_to_markdown(level: int) -> int:
    return max(2, min(level, 4))


def extract_page_structure(root: Tag, source_file: str, source_url: str | None = None) -> dict:
    container = find_page_container(root)
    title = find_page_title(root, source_file)
    page = {
        "source_file": source_file,
        "source_url": source_url,
        "marketplace": extract_marketplace(source_file),
        "title": title,
        "last_updated": find_last_updated(container),
        "blocks": [],
    }

    heading_stack: list[str] = [title]
    walk_content(root, page["blocks"], heading_stack)
    page["blocks"] = filter_leading_noise_blocks(page["blocks"])
    page["blocks"] = filter_tail_noise_blocks(page["blocks"])
    page["last_updated"] = None
    return page


def matches_any_pattern(text: str, patterns: tuple[re.Pattern[str], ...]) -> bool:
    return any(pattern.search(text) for pattern in patterns)


def is_tail_noise_block(block: dict) -> bool:
    text = normalize_semantic_text(block.get("text", ""))
    if not text:
        return False

    return matches_any_pattern(text, TAIL_UPDATED_PATTERNS) or matches_any_pattern(
        text,
        TAIL_DISCLAIMER_PATTERNS,
    )


def filter_tail_noise_blocks(blocks: list[dict]) -> list[dict]:
    filtered = list(blocks)
    while filtered and is_tail_noise_block(filtered[-1]):
        filtered.pop()
    return filtered


def is_leading_noise_block(block: dict) -> bool:
    text = normalize_semantic_text(block.get("text", ""))
    if not text:
        return False

    return matches_any_pattern(text, LEADING_NOISE_PATTERNS)


def filter_leading_noise_blocks(blocks: list[dict]) -> list[dict]:
    filtered = list(blocks)
    while filtered and is_leading_noise_block(filtered[0]):
        filtered.pop(0)
    return filtered


def extract_named_anchor(node: Tag) -> str | None:
    if node.name != "a":
        return None

    anchor = (node.get("name") or "").strip()
    return anchor or None


def is_anchor_only_tag(node: Tag) -> bool:
    return bool(extract_named_anchor(node)) and not normalize_text(node.get_text(" ", strip=True))


def current_block_anchor(anchor_state: dict[str, str | None]) -> str | None:
    return anchor_state.get("current_anchor")


def update_anchor_state_from_node(node: Tag, anchor_state: dict[str, str | None]) -> str | None:
    anchor = extract_named_anchor(node)
    if anchor:
        anchor_state["current_anchor"] = anchor
    return anchor


def walk_content(
    node: Tag,
    blocks: list[dict],
    heading_stack: list[str],
    anchor_state: dict[str, str | None] | None = None,
) -> None:
    if anchor_state is None:
        anchor_state = {"current_anchor": None}

    for child in node.children:
        if isinstance(child, NavigableString):
            continue
        if not isinstance(child, Tag):
            continue
        if should_drop_node(child):
            continue

        update_anchor_state_from_node(child, anchor_state)
        if is_anchor_only_tag(child):
            continue

        if child.name in {"h2", "h3", "h4"}:
            text = normalize_text(child.get_text(" ", strip=True))
            if not text or text in TITLE_EXCLUSIONS:
                continue
            if text == heading_stack[0] and len(heading_stack) == 1 and not blocks:
                continue

            level = heading_level_to_markdown(int(child.name[1]))
            relative_depth = max(0, level - 2)
            heading_stack[:] = heading_stack[: relative_depth + 1]
            heading_stack.append(text)
            blocks.append(
                {
                    "type": "heading",
                    "level": level,
                    "text": text,
                    "heading_path": " > ".join(heading_stack),
                    "anchor": current_block_anchor(anchor_state),
                }
            )
            continue

        if child.name == "p":
            text = normalize_text(child.get_text(" ", strip=True))
            if text:
                blocks.append(
                    {
                        "type": "paragraph",
                        "text": text,
                        "heading_path": " > ".join(heading_stack),
                        "anchor": current_block_anchor(anchor_state),
                    }
                )
            continue

        if child.name in {"ul", "ol"}:
            blocks.extend(parse_list(child, " > ".join(heading_stack), anchor_state=anchor_state))
            continue

        if child.name == "table":
            table_text = render_table(child)
            if table_text:
                blocks.append(
                    {
                        "type": "table",
                        "text": table_text,
                        "heading_path": " > ".join(heading_stack),
                        "anchor": current_block_anchor(anchor_state),
                    }
                )
            continue

        walk_content(child, blocks, heading_stack, anchor_state)


def parse_list(
    list_tag: Tag,
    heading_path: str,
    indent: int = 0,
    anchor_state: dict[str, str | None] | None = None,
) -> list[dict]:
    if anchor_state is None:
        anchor_state = {"current_anchor": None}

    blocks: list[dict] = []

    for child in list_tag.children:
        if isinstance(child, NavigableString):
            continue
        if not isinstance(child, Tag):
            continue

        update_anchor_state_from_node(child, anchor_state)
        if is_anchor_only_tag(child):
            continue
        if child.name != "li":
            continue

        li = child
        for nested_child in li.children:
            if not isinstance(nested_child, Tag):
                continue
            update_anchor_state_from_node(nested_child, anchor_state)

        li_clone = clone_tag(li)
        for nested in li_clone.find_all(["ul", "ol"]):
            nested.decompose()

        text = normalize_text(li_clone.get_text(" ", strip=True))
        if text:
            blocks.append(
                {
                    "type": "list_item",
                    "text": text,
                    "indent": indent,
                    "heading_path": heading_path,
                    "anchor": current_block_anchor(anchor_state),
                }
            )

        for nested_list in li.find_all(["ul", "ol"], recursive=False):
            blocks.extend(parse_list(nested_list, heading_path, indent + 1, anchor_state))

    return blocks


def clone_tag(node: Tag) -> Tag:
    try:
        clone = BeautifulSoup(str(node), "lxml")
    except FeatureNotFound:
        clone = BeautifulSoup(str(node), "html.parser")
    return clone.find(node.name)


def render_table(table: Tag) -> str:
    rows = []
    for tr in table.find_all("tr"):
        cells = [normalize_text(cell.get_text(" ", strip=True)) for cell in tr.find_all(["th", "td"])]
        cells = [cell for cell in cells if cell]
        if cells:
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def render_extracted_markdown(page: dict) -> str:
    lines = [
        "---",
        f"source_file: {page['source_file']}",
        f"title: {page['title']}",
    ]
    if page.get("source_url"):
        lines.append(f"source_url: {page['source_url']}")
    if page.get("marketplace"):
        lines.append(f"marketplace: {page['marketplace']}")
    lines.extend(["---", "", f"# {page['title']}", ""])

    previous_type = None
    for block in page["blocks"]:
        if block["type"] == "heading":
            if lines and lines[-1] != "":
                lines.append("")
            lines.append(f"{'#' * block['level']} {block['text']}")
            lines.append("")
        elif block["type"] == "paragraph":
            if previous_type == "list_item" and lines[-1] != "":
                lines.append("")
            lines.append(block["text"])
            lines.append("")
        elif block["type"] == "list_item":
            lines.append(f"{'  ' * block.get('indent', 0)}- {block['text']}")
        elif block["type"] == "table":
            if previous_type == "list_item" and lines[-1] != "":
                lines.append("")
            lines.append(block["text"])
            lines.append("")
        previous_type = block["type"]

    return "\n".join(trim_trailing_blank_lines(lines)) + "\n"


def render_page_markdown(page: dict) -> str:
    return render_extracted_markdown(page)


def trim_trailing_blank_lines(lines: list[str]) -> list[str]:
    trimmed = copy.copy(lines)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def render_blocks_for_chunk_output(blocks: list[dict]) -> str:
    lines: list[str] = []
    previous_type = None

    for block in blocks:
        if block["type"] == "paragraph":
            if previous_type == "list_item" and lines and lines[-1] != "":
                lines.append("")
            lines.append(block["text"])
            lines.append("")
        elif block["type"] == "list_item":
            lines.append(f"{'  ' * block.get('indent', 0)}- {block['text']}")
        elif block["type"] == "table":
            if previous_type == "list_item" and lines and lines[-1] != "":
                lines.append("")
            lines.append(block["text"])
            lines.append("")
        previous_type = block["type"]

    return "\n".join(trim_trailing_blank_lines(lines)).strip()


def split_blocks_for_chunk_output(
    blocks: list[dict],
    soft_max: int = 600,
    hard_max: int = 900,
    preferred_min_items: int = 4,
    preferred_max_items: int = 6,
) -> list[list[dict]]:
    if not blocks:
        return []

    if not any(block["type"] == "list_item" for block in blocks):
        return [blocks]

    chunks: list[list[dict]] = []
    current: list[dict] = []
    current_item_count = 0

    def flush() -> None:
        nonlocal current, current_item_count
        if current:
            chunks.append(current)
        current = []
        current_item_count = 0

    for block in blocks:
        if block["type"] != "list_item":
            current.append(block)
            continue

        item_length = len(render_blocks_for_chunk_output([block]))
        if item_length > hard_max:
            flush()
            chunks.append([block])
            continue

        candidate = current + [block]
        candidate_length = len(render_blocks_for_chunk_output(candidate))
        should_flush = False

        if current_item_count >= preferred_max_items:
            should_flush = True
        elif current and candidate_length > soft_max and current_item_count >= max(1, preferred_min_items - 1):
            should_flush = True
        elif current and candidate_length > hard_max:
            should_flush = True

        if should_flush:
            flush()

        current.append(block)
        current_item_count += 1

    flush()
    return chunks if chunks else [blocks]


def build_semantic_sections(
    blocks: list[dict],
    intro_heading: str | None = None,
) -> list[dict]:
    sections: list[dict] = []
    current_heading = intro_heading
    current_anchor = None
    current_blocks: list[dict] = []

    def flush() -> None:
        nonlocal current_blocks
        if current_blocks:
            sections.append(
                {
                    "heading": current_heading,
                    "anchor": current_anchor,
                    "blocks": current_blocks,
                }
            )
        current_blocks = []

    for block in blocks:
        if block["type"] == "heading":
            flush()
            current_heading = block["text"]
            current_anchor = block.get("anchor")
            continue

        current_blocks.append(block)

    flush()
    return sections


def find_first_block_anchor(blocks: list[dict]) -> str | None:
    for block in blocks:
        anchor = block.get("anchor")
        if anchor:
            return anchor
    return None


def resolve_chunk_anchor(
    section_anchor: str | None,
    chunk_blocks: list[dict],
    is_first_chunk: bool,
) -> str | None:
    block_anchor = find_first_block_anchor(chunk_blocks)
    if is_first_chunk:
        return section_anchor or block_anchor
    return block_anchor or section_anchor


def build_semantic_chunk_entries(
    page: dict,
    soft_max: int = 600,
    hard_max: int = 900,
    preferred_min_items: int = 4,
    preferred_max_items: int = 6,
    intro_heading_fallback: bool = True,
) -> list[dict]:
    intro_heading = page.get("title") if intro_heading_fallback else None
    sections = build_semantic_sections(page.get("blocks", []), intro_heading=intro_heading)
    chunks: list[dict] = []

    for section in sections:
        split_chunks = split_blocks_for_chunk_output(
            section["blocks"],
            soft_max=soft_max,
            hard_max=hard_max,
            preferred_min_items=preferred_min_items,
            preferred_max_items=preferred_max_items,
        )
        for chunk_index, chunk_blocks in enumerate(split_chunks):
            chunks.append(
                {
                    "heading": section["heading"],
                    "anchor": resolve_chunk_anchor(
                        section.get("anchor"),
                        chunk_blocks,
                        is_first_chunk=chunk_index == 0,
                    ),
                    "blocks": chunk_blocks,
                }
            )

    return chunks


def build_chunk_source_url(page_source_url: str | None, anchor: str | None) -> str | None:
    if not page_source_url:
        return None
    if not anchor:
        return page_source_url
    return f"{page_source_url.split('#', 1)[0]}#{anchor}"


def render_semantic_chunks_markdown(
    page: dict,
    soft_max: int = 600,
    hard_max: int = 900,
    preferred_min_items: int = 4,
    preferred_max_items: int = 6,
    intro_heading_fallback: bool = True,
) -> str:
    chunks = build_semantic_chunk_entries(
        page,
        soft_max=soft_max,
        hard_max=hard_max,
        preferred_min_items=preferred_min_items,
        preferred_max_items=preferred_max_items,
        intro_heading_fallback=intro_heading_fallback,
    )

    lines = [f"# {page['title']}", ""]

    for index, chunk in enumerate(chunks, start=1):
        lines.append(f"## Chunk {index}")
        chunk_source_url = build_chunk_source_url(page.get("source_url"), chunk.get("anchor"))
        if chunk_source_url:
            lines.append(f"source_url: {chunk_source_url}")
        lines.append(f"source_file: {page['source_file']}")
        lines.append("")

        if chunk.get("heading"):
            lines.append(f"### {chunk['heading']}")
            lines.append("")

        body = render_blocks_for_chunk_output(chunk["blocks"])
        if body:
            lines.append(body)
            lines.append("")

    return "\n".join(trim_trailing_blank_lines(lines)) + "\n"


def build_chunk_id(page: dict, chunk_number: int) -> str:
    marketplace = page.get("marketplace") or "unknown"
    page_id = extract_page_id(page.get("source_file", "")) or Path(page.get("source_file", "page")).stem
    return f"{marketplace}_{page_id}_c{chunk_number:03d}"


def truncate_text(value: str, max_chars: int = 240) -> str:
    normalized = normalize_text(value)
    if len(normalized) <= max_chars:
        return normalized
    return normalized[: max_chars - 3].rstrip() + "..."


def first_heading_path(chunk_blocks: list[dict], fallback: str) -> str:
    for block in chunk_blocks:
        heading_path = normalize_text(block.get("heading_path", ""))
        if heading_path:
            return heading_path
    return fallback


def collect_semantic_terms(text: str, rules: tuple[tuple[re.Pattern[str], str], ...]) -> list[str]:
    terms: list[str] = []
    for pattern, term in rules:
        if pattern.search(text):
            terms.append(term)
    return unique_preserving_order(terms)


def build_seed_terms_for_chunk(
    chunk_text: str,
    heading: str,
    page_entities: list[str],
) -> dict:
    combined_text = f"{heading} {chunk_text}"
    category_terms = build_term_candidates(combined_text, limit=10)
    if page_entities:
        category_terms = unique_preserving_order([*category_terms, *page_entities[:5]])[:12]

    risk_terms = collect_semantic_terms(combined_text, RISK_TERM_RULES)
    scenario_terms = collect_semantic_terms(combined_text, SCENARIO_TERM_RULES)

    variant_terms: list[str] = []
    for term in [*risk_terms, *scenario_terms]:
        variant_terms.extend(SEMANTIC_VARIANTS.get(term, []))
    if "compliance" in normalize_semantic_text(combined_text):
        variant_terms.append("regulatory compliance")
    variant_terms = unique_preserving_order(variant_terms)

    return {
        "category_terms": category_terms,
        "risk_terms": risk_terms,
        "scenario_terms": scenario_terms,
        "variant_terms": variant_terms,
    }


def build_index_seed_for_page(page: dict) -> dict:
    chunks = build_semantic_chunk_entries(page)
    page_entities = extract_global_entities(page)

    page_profile = {
        "source_file": page.get("source_file"),
        "source_url": page.get("source_url"),
        "marketplace": page.get("marketplace"),
        "page_title": page.get("title"),
        "page_summary": build_page_summary(page),
        "global_entities": page_entities,
    }

    chunk_packets: list[dict] = []
    for index, chunk in enumerate(chunks, start=1):
        chunk_text = render_blocks_for_chunk_output(chunk.get("blocks", []))
        heading = chunk.get("heading") or page.get("title")
        fallback_path = normalize_text(" > ".join([page.get("title", ""), heading or ""]))
        chunk_packets.append(
            {
                "chunk_id": build_chunk_id(page, index),
                "heading": heading,
                "heading_path": first_heading_path(chunk.get("blocks", []), fallback=fallback_path),
                "chunk_text": chunk_text,
                "chunk_source_url": build_chunk_source_url(page.get("source_url"), chunk.get("anchor")),
                "prev_context": "",
                "next_context": "",
            }
        )

    for index, packet in enumerate(chunk_packets):
        if index > 0:
            packet["prev_context"] = truncate_text(chunk_packets[index - 1]["chunk_text"])
        if index + 1 < len(chunk_packets):
            packet["next_context"] = truncate_text(chunk_packets[index + 1]["chunk_text"])

    seed_terms: list[dict] = []
    for packet in chunk_packets:
        terms = build_seed_terms_for_chunk(
            chunk_text=packet["chunk_text"],
            heading=packet.get("heading") or "",
            page_entities=page_entities,
        )
        seed_terms.append({"chunk_id": packet["chunk_id"], **terms})

    return {
        "page_profile": page_profile,
        "chunk_packet": chunk_packets,
        "seed_terms": seed_terms,
    }


def render_single_block_for_llm_packet(block: dict) -> str:
    block_type = block.get("type")
    if block_type == "list_item":
        return f"{'  ' * block.get('indent', 0)}- {block.get('text', '')}".rstrip()
    return normalize_text(block.get("text", ""))


def build_doc_id(page: dict) -> str:
    marketplace = page.get("marketplace") or "unknown"
    page_id = extract_page_id(page.get("source_file", "")) or Path(page.get("source_file", "page")).stem
    return f"{marketplace}_{page_id}"


def build_llm_chunk_packet_for_page(page: dict) -> list[dict]:
    packets: list[dict] = []
    current_heading = page.get("title")

    for index, block in enumerate(page.get("blocks", []), start=1):
        block_type = block.get("type")
        if block_type == "heading":
            current_heading = block.get("text") or current_heading
            continue

        text = render_single_block_for_llm_packet(block)
        if not text:
            continue

        packet = {
            "chunk_id": f"{build_doc_id(page)}_u{index:04d}",
            "heading": current_heading or page.get("title"),
            "heading_path": block.get("heading_path") or normalize_text(
                " > ".join([page.get("title", ""), current_heading or ""])
            ),
            "chunk_text": text,
            "chunk_source_url": build_chunk_source_url(page.get("source_url"), block.get("anchor")),
            "prev_context": "",
            "next_context": "",
            "packet_kind": block_type,
            "packet_order": index,
            "bucket_hints": infer_regulation_bucket_hints(
                current_heading or "",
                block.get("heading_path", ""),
                text,
            ),
        }
        packets.append(packet)

    for idx, packet in enumerate(packets):
        if idx > 0:
            packet["prev_context"] = truncate_text(packets[idx - 1]["chunk_text"], max_chars=220)
        if idx + 1 < len(packets):
            packet["next_context"] = truncate_text(packets[idx + 1]["chunk_text"], max_chars=220)

    return packets


def build_llm_seed_terms_for_packets(
    page: dict,
    chunk_packets: list[dict],
    page_entities: list[str],
) -> list[dict]:
    seeds: list[dict] = []
    for packet in chunk_packets:
        terms = build_seed_terms_for_chunk(
            chunk_text=packet.get("chunk_text", ""),
            heading=packet.get("heading") or page.get("title", ""),
            page_entities=page_entities,
        )
        seeds.append(
            {
                "chunk_id": packet["chunk_id"],
                "seed_lang": "source+en",
                **terms,
            }
        )
    return seeds


def build_llm_doc_packet_for_page(
    page: dict,
    subagent_model: str | None = None,
    fallback_subagent_models: Any = None,
) -> dict:
    resolved_subagent_model = normalize_subagent_model(subagent_model)
    resolved_fallback_models = normalize_subagent_model_list(
        fallback_subagent_models,
        primary_model=resolved_subagent_model,
    )
    page_entities = extract_global_entities(page)
    page_profile = {
        "doc_id": build_doc_id(page),
        "source_file": page.get("source_file"),
        "source_url": page.get("source_url"),
        "marketplace": page.get("marketplace"),
        "page_title": page.get("title"),
        "page_summary": build_page_summary(page),
        "global_entities": page_entities,
        "bucket_hints": infer_regulation_bucket_hints(page.get("title", ""), build_page_summary(page)),
    }
    chunk_packets = build_llm_chunk_packet_for_page(page)
    seed_terms = build_llm_seed_terms_for_packets(page, chunk_packets, page_entities)

    return {
        "doc_id": page_profile["doc_id"],
        "generated_at": utc_now_iso(),
        "page_profile": page_profile,
        "chunk_packet": chunk_packets,
        "seed_terms": seed_terms,
        "orchestration": {
            "stage": "A",
            "subagent_scope": "document",
            "prompt_template_path": stage_a_prompt_template_path(),
            "subagent_model": resolved_subagent_model,
            "fallback_subagent_models": resolved_fallback_models,
        },
        "contract": {
            "required_input": ["page_profile", "chunk_packet", "seed_terms"],
            "required_output": {
                "doc_id": "string",
                "semantic_chunks": [
                    {
                        "chunk_id": "string",
                        "heading": "string",
                        "heading_path": "string",
                        "chunk_text": "string",
                        "chunk_source_url": "string",
                        "prev_context": "string",
                        "next_context": "string",
                        "screening_relevance_score": "integer(0-10)",
                        "screening_relevance_reason": "string",
                    }
                ],
            },
            "notes": [
                "阶段 A 只做语义分段，不输出 final_terms/final_title。",
                "每个 semantic chunk 需要输出禁售筛查相关性分（0-10），用于评估是否值得纳入向量库。",
                "阶段 B 再逐 chunk 输出 final_terms。",
            ],
        },
    }


def write_jsonl_records(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def infer_orchestration_stage(record: dict) -> str:
    orchestration = record.get("orchestration")
    if isinstance(orchestration, dict):
        stage = normalize_text(str(orchestration.get("stage", "")))
        if stage:
            return stage
    if isinstance(record.get("chunk_packet"), list):
        return "A"
    if isinstance(record.get("chunk_packet"), dict):
        return "B"
    return "unknown"


def default_required_output_for_stage(stage: str) -> dict:
    if stage == "A":
        return {
            "doc_id": "string",
            "semantic_chunks": [
                {
                    "chunk_id": "string",
                    "heading": "string",
                    "heading_path": "string",
                    "chunk_text": "string",
                    "chunk_source_url": "string",
                    "prev_context": "string",
                    "next_context": "string",
                    "screening_relevance_score": "integer(0-10)",
                    "screening_relevance_reason": "string",
                }
            ],
        }
    if stage == "B":
        return {
            "doc_id": "string",
            "chunk_id": "string",
            "final_terms": ["string"],
        }
    return {}


def preview_filename_for_record(record: dict, index: int) -> str | None:
    stage = infer_orchestration_stage(record)
    doc_id = normalize_text(str(record.get("doc_id", ""))) or f"record-{index:04d}"
    chunk_id = normalize_text(str(record.get("chunk_id", "")))
    if stage == "A":
        filename = f"{index:04d}-A-{doc_id}.md"
    elif stage == "B":
        if not chunk_id and isinstance(record.get("chunk_packet"), dict):
            chunk_id = normalize_text(str(record["chunk_packet"].get("chunk_id", "")))
        filename = f"{index:04d}-B-{doc_id}-{chunk_id or 'chunk'}.md"
    else:
        return None
    return re.sub(r"[^A-Za-z0-9._-]+", "_", filename)


def runner_input_filename_for_record(record: dict, index: int) -> str | None:
    preview_name = preview_filename_for_record(record, index)
    if not preview_name:
        return None
    stem = Path(preview_name).stem
    return f"{stem}.runner.md"


def build_task_manifest(records: list[dict], input_path: Path, output_root: Path) -> dict:
    tasks: list[dict] = []
    stage_counts: Counter[str] = Counter()
    preview_dir = default_prompt_preview_dir(input_path, output_root)
    runner_input_dir = default_runner_input_dir(input_path, output_root)

    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            continue
        stage = infer_orchestration_stage(record)
        stage_counts[stage] += 1

        orchestration = record.get("orchestration", {}) if isinstance(record.get("orchestration"), dict) else {}
        prompt_template_path = normalize_text(str(orchestration.get("prompt_template_path", "")))
        if not prompt_template_path:
            prompt_template_path = stage_a_prompt_template_path() if stage == "A" else stage_b_prompt_template_path() if stage == "B" else ""
        doc_id = normalize_text(str(record.get("doc_id", "")))
        chunk_id = normalize_text(str(record.get("chunk_id", "")))
        chunk_packet = record.get("chunk_packet")
        if not chunk_id and isinstance(chunk_packet, dict):
            chunk_id = normalize_text(str(chunk_packet.get("chunk_id", "")))
        preview_name = preview_filename_for_record(record, index)
        runner_input_name = runner_input_filename_for_record(record, index)

        task = {
            "task_id": f"{stage.lower()}-task-{index:04d}",
            "stage": stage,
            "subagent_scope": normalize_text(str(orchestration.get("subagent_scope", ""))) or (
                "document" if stage == "A" else "chunk" if stage == "B" else ""
            ),
            "subagent_model": normalize_subagent_model(orchestration.get("subagent_model")),
            "fallback_subagent_models": normalize_subagent_model_list(
                orchestration.get("fallback_subagent_models"),
                primary_model=orchestration.get("subagent_model"),
            ),
            "prompt_template_path": prompt_template_path,
            "preview_markdown_path": str(preview_dir / preview_name) if preview_name else "",
            "runner_input_path": str(runner_input_dir / runner_input_name) if runner_input_name else "",
            "input_jsonl_path": str(input_path),
            "input_record_number": index,
            "doc_id": doc_id,
            "chunk_id": chunk_id,
            "packet_count": len(record.get("chunk_packet", [])) if isinstance(record.get("chunk_packet"), list) else 1 if isinstance(record.get("chunk_packet"), dict) else 0,
            "expected_output": default_required_output_for_stage(stage),
        }
        tasks.append(task)

    return {
        "generated_at": utc_now_iso(),
        "source_jsonl_path": str(input_path),
        "task_count": len(tasks),
        "stage_counts": dict(stage_counts),
        "tasks": tasks,
    }


def render_stage_a_prompt_preview(record: dict) -> str:
    page_profile = record.get("page_profile", {}) if isinstance(record.get("page_profile"), dict) else {}
    doc_id = normalize_text(str(record.get("doc_id", ""))) or normalize_text(str(page_profile.get("doc_id", "")))
    title = normalize_text(str(page_profile.get("page_title", ""))) or doc_id or "Untitled"
    marketplace = normalize_text(str(page_profile.get("marketplace", "")))
    source_file = normalize_text(str(page_profile.get("source_file", "")))
    chunk_packet = record.get("chunk_packet", [])

    lines = [
        f"# {title}",
        "",
        "## Minimal Metadata",
        f"- doc_id: {doc_id}",
        f"- marketplace: {marketplace}",
        f"- source_file: {source_file}",
        "",
        "## Regulation Source Markdown",
        "",
    ]

    last_heading = ""
    for packet in chunk_packet if isinstance(chunk_packet, list) else []:
        if not isinstance(packet, dict):
            continue
        heading = normalize_text(str(packet.get("heading", "")))
        chunk_text = normalize_text(str(packet.get("chunk_text", "")))
        if not chunk_text:
            continue
        if heading and heading != title and heading != last_heading:
            lines.extend([f"## {heading}", ""])
            last_heading = heading
        lines.extend([chunk_text, ""])

    lines.extend(
        [
            "## Notes",
            "- This preview is markdown-first for subagent execution.",
            "- Transport-only URL and adjacency metadata are intentionally omitted.",
        ]
    )
    return "\n".join(lines).strip() + "\n"


def render_stage_b_prompt_preview(record: dict) -> str:
    page_profile = record.get("page_profile", {}) if isinstance(record.get("page_profile"), dict) else {}
    chunk_packet = record.get("chunk_packet", {}) if isinstance(record.get("chunk_packet"), dict) else {}
    seed_terms = record.get("seed_terms", {}) if isinstance(record.get("seed_terms"), dict) else {}
    doc_id = normalize_text(str(record.get("doc_id", ""))) or normalize_text(str(page_profile.get("doc_id", "")))
    page_title = normalize_text(str(page_profile.get("page_title", ""))) or doc_id or "Untitled"
    bucket_hints = unique_preserving_order([str(value) for value in chunk_packet.get("bucket_hints", [])]) if isinstance(chunk_packet.get("bucket_hints"), list) else []
    screening_relevance_score = normalize_screening_relevance_score(chunk_packet.get("screening_relevance_score"))
    screening_relevance_reason = normalize_text(str(chunk_packet.get("screening_relevance_reason", "")))

    def join_terms(key: str) -> str:
        values = seed_terms.get(key, [])
        if not isinstance(values, list):
            return ""
        return " | ".join(unique_preserving_order([str(value) for value in values]))

    lines = [
        f"# {page_title}",
        "",
        "## Minimal Metadata",
        f"- doc_id: {doc_id}",
        f"- chunk_id: {normalize_text(str(chunk_packet.get('chunk_id', '')))}",
        f"- heading: {normalize_text(str(chunk_packet.get('heading', '')))}",
        f"- heading_path: {normalize_text(str(chunk_packet.get('heading_path', '')))}",
        f"- bucket_hints: {' | '.join(bucket_hints)}",
        f"- screening_relevance_score: {'' if screening_relevance_score is None else screening_relevance_score}",
        f"- screening_relevance_reason: {screening_relevance_reason}",
        "",
        "## Semantic Chunk",
        "",
        normalize_text(str(chunk_packet.get("chunk_text", ""))),
        "",
        "## Seed Terms",
        f"- category_terms: {join_terms('category_terms')}",
        f"- risk_terms: {join_terms('risk_terms')}",
        f"- scenario_terms: {join_terms('scenario_terms')}",
        f"- variant_terms: {join_terms('variant_terms')}",
        "",
        "## Notes",
        "- This preview is markdown-first for subagent execution.",
        "- Transport-only URL and adjacency metadata are intentionally omitted.",
    ]
    return "\n".join(lines).strip() + "\n"


def write_prompt_previews(records: list[dict], input_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            continue
        stage = infer_orchestration_stage(record)
        if stage == "A":
            content = render_stage_a_prompt_preview(record)
        elif stage == "B":
            content = render_stage_b_prompt_preview(record)
        else:
            continue
        safe_name = preview_filename_for_record(record, index)
        if not safe_name:
            continue
        (output_dir / safe_name).write_text(content, encoding="utf-8")
    return output_dir


def build_runner_prompt_content(template_path: Path, preview_path: Path) -> str:
    template_text = template_path.read_text(encoding="utf-8").strip()
    preview_text = preview_path.read_text(encoding="utf-8").strip()
    return (
        "## System Prompt Template\n\n"
        f"{template_text}\n\n"
        "---\n\n"
        "## Input Context\n\n"
        "The content below is the markdown-first execution packet. Apply the system prompt template above to this input context only.\n\n"
        "## Execution Packet\n\n"
        f"{preview_text}\n"
    )


def write_runner_inputs(
    records: list[dict],
    input_path: Path,
    preview_dir: Path,
    runner_dir: Path,
) -> Path:
    if not preview_dir.exists():
        write_prompt_previews(records, input_path, preview_dir)

    runner_dir.mkdir(parents=True, exist_ok=True)
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            continue
        orchestration = record.get("orchestration", {}) if isinstance(record.get("orchestration"), dict) else {}
        stage = infer_orchestration_stage(record)
        prompt_template_path = normalize_text(str(orchestration.get("prompt_template_path", "")))
        if not prompt_template_path:
            prompt_template_path = stage_a_prompt_template_path() if stage == "A" else stage_b_prompt_template_path() if stage == "B" else ""
        if not prompt_template_path:
            continue
        preview_name = preview_filename_for_record(record, index)
        runner_name = runner_input_filename_for_record(record, index)
        if not preview_name or not runner_name:
            continue
        preview_path = preview_dir / preview_name
        if not preview_path.exists():
            if stage == "A":
                preview_path.write_text(render_stage_a_prompt_preview(record), encoding="utf-8")
            elif stage == "B":
                preview_path.write_text(render_stage_b_prompt_preview(record), encoding="utf-8")
        content = build_runner_prompt_content(Path(prompt_template_path), preview_path)
        primary_model = normalize_subagent_model(orchestration.get("subagent_model"))
        fallback_models = normalize_subagent_model_list(
            orchestration.get("fallback_subagent_models"),
            primary_model=primary_model,
        )
        if fallback_models:
            content += (
                "\n## Runner Hints\n\n"
                f"- primary_subagent_model: {primary_model}\n"
                f"- fallback_subagent_models: {' | '.join(fallback_models)}\n"
                "- If the runner receives `Selected model is at capacity. Please try a different model.`, retry with the fallback models in order.\n"
            )
        (runner_dir / runner_name).write_text(content, encoding="utf-8")
    return runner_dir


def load_json_or_jsonl(path: Path) -> list[dict]:
    if path.suffix.lower() == ".jsonl":
        records: list[dict] = []
        for line in path.read_text(encoding="utf-8").splitlines():
            text = line.strip()
            if not text:
                continue
            payload = json.loads(text)
            if isinstance(payload, dict):
                records.append(payload)
        return records

    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    if isinstance(payload.get("documents"), list):
        return [item for item in payload["documents"] if isinstance(item, dict)]
    if isinstance(payload.get("results"), list):
        return [item for item in payload["results"] if isinstance(item, dict)]
    if "doc_id" in payload or "chunk_terms" in payload or "semantic_chunks" in payload:
        return [payload]
    return []


def normalize_page_profile(profile: Any) -> dict:
    if not isinstance(profile, dict):
        return {}

    normalized: dict[str, Any] = {}
    for key in ("doc_id", "source_file", "source_url", "marketplace", "page_title", "page_summary"):
        value = normalize_text(str(profile.get(key, "")))
        if value:
            normalized[key] = value

    global_entities = normalize_term_list(profile.get("global_entities"))
    if global_entities:
        normalized["global_entities"] = global_entities

    bucket_hints = (
        unique_preserving_order([str(value) for value in profile.get("bucket_hints", [])])
        if isinstance(profile.get("bucket_hints"), list)
        else []
    )
    if bucket_hints:
        normalized["bucket_hints"] = bucket_hints

    for key, value in profile.items():
        if key in normalized:
            continue
        if isinstance(value, str):
            text = normalize_text(value)
            if text:
                normalized[key] = text
        elif value not in (None, "", [], {}):
            normalized[key] = copy.deepcopy(value)
    return normalized


def merge_page_profiles(primary: Any, fallback: Any) -> dict:
    primary_profile = normalize_page_profile(primary)
    fallback_profile = normalize_page_profile(fallback)
    if not primary_profile:
        return fallback_profile
    if not fallback_profile:
        return primary_profile

    merged = copy.deepcopy(primary_profile)
    for key, value in fallback_profile.items():
        if key in {"global_entities", "bucket_hints"}:
            existing = merged.get(key, [])
            if not isinstance(existing, list):
                existing = []
            merged[key] = unique_preserving_order([*existing, *value]) if isinstance(value, list) else existing
            continue

        current = merged.get(key)
        if isinstance(current, str):
            current = normalize_text(current)
        if current in (None, "", [], {}):
            merged[key] = copy.deepcopy(value)
    return merged


def build_page_profile_index(*paths: Path) -> dict[str, dict]:
    search_dirs: list[Path] = []
    seen_dirs: set[Path] = set()
    for raw_path in paths:
        if raw_path is None:
            continue
        path = Path(raw_path)
        directory = path if path.is_dir() else path.parent
        if directory in seen_dirs or not directory.exists():
            continue
        seen_dirs.add(directory)
        search_dirs.append(directory)

    candidate_files: list[Path] = []
    seen_files: set[Path] = set()
    for directory in search_dirs:
        for pattern in ("*.jsonl", "*.json"):
            for file_path in sorted(directory.glob(pattern)):
                if file_path in seen_files or not file_path.is_file():
                    continue
                seen_files.add(file_path)
                candidate_files.append(file_path)

    profile_index: dict[str, dict] = {}
    for file_path in candidate_files:
        try:
            records = load_json_or_jsonl(file_path)
        except Exception:
            continue
        for record in records:
            if not isinstance(record, dict):
                continue
            page_profile = normalize_page_profile(record.get("page_profile"))
            doc_id = normalize_text(str(record.get("doc_id", ""))) or normalize_text(str(page_profile.get("doc_id", "")))
            if not doc_id or not page_profile:
                continue
            page_profile.setdefault("doc_id", doc_id)
            profile_index[doc_id] = merge_page_profiles(profile_index.get(doc_id, {}), page_profile)
    return profile_index


def normalize_chunk_terms_record(entry: Any) -> dict | None:
    if not isinstance(entry, dict):
        return None
    chunk_id = normalize_text(str(entry.get("chunk_id", "")))
    if not chunk_id:
        return None

    final_terms = normalize_term_list(entry.get("final_terms"))
    if not final_terms:
        return None
    return {"chunk_id": chunk_id, "final_terms": final_terms}


def normalize_doc_result_record(record: dict, page_profile_index: dict[str, dict] | None = None) -> dict | None:
    doc_id = normalize_text(str(record.get("doc_id", "")))
    page_profile = record.get("page_profile", {}) if isinstance(record.get("page_profile"), dict) else {}

    chunk_terms_raw = record.get("chunk_terms")
    if chunk_terms_raw is None:
        chunk_terms_raw = record.get("chunks")
    if chunk_terms_raw is None and record.get("chunk_id") and record.get("final_terms"):
        chunk_terms_raw = [{"chunk_id": record.get("chunk_id"), "final_terms": record.get("final_terms")}]
    if not isinstance(chunk_terms_raw, list):
        return None

    chunk_terms: list[dict] = []
    for entry in chunk_terms_raw:
        normalized = normalize_chunk_terms_record(entry)
        if normalized:
            chunk_terms.append(normalized)

    if not chunk_terms:
        return None
    if not doc_id:
        doc_id = normalize_text(str(page_profile.get("doc_id", ""))) or "unknown_doc"
    page_profile = merge_page_profiles(page_profile, (page_profile_index or {}).get(doc_id, {}))
    if page_profile:
        page_profile.setdefault("doc_id", doc_id)

    return {
        "doc_id": doc_id,
        "page_profile": page_profile,
        "chunk_terms": chunk_terms,
    }


def normalize_semantic_chunk_record(entry: Any) -> dict | None:
    if not isinstance(entry, dict):
        return None
    chunk_id = normalize_text(str(entry.get("chunk_id", "")))
    chunk_text = normalize_text(str(entry.get("chunk_text", "")))
    if not chunk_id or not chunk_text:
        return None

    bucket_hints = (
        unique_preserving_order([str(value) for value in entry.get("bucket_hints", [])])
        if isinstance(entry.get("bucket_hints"), list)
        else []
    )
    explicit_score = normalize_screening_relevance_score(entry.get("screening_relevance_score"))
    reason = normalize_text(str(entry.get("screening_relevance_reason", "")))

    return {
        "chunk_id": chunk_id,
        "heading": normalize_text(str(entry.get("heading", ""))),
        "heading_path": normalize_text(str(entry.get("heading_path", ""))),
        "chunk_text": chunk_text,
        "chunk_source_url": normalize_text(str(entry.get("chunk_source_url", ""))),
        "prev_context": normalize_text(str(entry.get("prev_context", ""))),
        "next_context": normalize_text(str(entry.get("next_context", ""))),
        "bucket_hints": bucket_hints,
        "screening_relevance_score": explicit_score,
        "screening_relevance_reason": reason,
    }


def normalize_semantic_doc_record(record: dict, page_profile_index: dict[str, dict] | None = None) -> dict | None:
    doc_id = normalize_text(str(record.get("doc_id", "")))
    page_profile = record.get("page_profile", {}) if isinstance(record.get("page_profile"), dict) else {}

    semantic_chunks_raw = record.get("semantic_chunks")
    if semantic_chunks_raw is None:
        semantic_chunks_raw = record.get("chunks")
    if semantic_chunks_raw is None and record.get("chunk_id") and record.get("chunk_text"):
        semantic_chunks_raw = [record]
    if not isinstance(semantic_chunks_raw, list):
        return None

    semantic_chunks: list[dict] = []
    for entry in semantic_chunks_raw:
        normalized = normalize_semantic_chunk_record(entry)
        if normalized:
            semantic_chunks.append(normalized)

    if not semantic_chunks:
        return None
    if not doc_id:
        doc_id = normalize_text(str(page_profile.get("doc_id", ""))) or "unknown_doc"
    page_profile = merge_page_profiles(page_profile, (page_profile_index or {}).get(doc_id, {}))
    if page_profile:
        page_profile.setdefault("doc_id", doc_id)

    return {
        "doc_id": doc_id,
        "page_profile": page_profile,
        "semantic_chunks": semantic_chunks,
    }


def merge_llm_doc_results(records: list[dict], page_profile_index: dict[str, dict] | None = None) -> dict:
    documents: dict[str, dict] = {}

    for record in records:
        normalized = normalize_doc_result_record(record, page_profile_index=page_profile_index)
        if not normalized:
            continue

        doc_id = normalized["doc_id"]
        slot = documents.setdefault(
            doc_id,
            {
                "doc_id": doc_id,
                "page_profile": normalized.get("page_profile", {}),
                "chunk_terms": [],
            },
        )
        slot["page_profile"] = merge_page_profiles(slot.get("page_profile", {}), normalized.get("page_profile", {}))
        slot["chunk_terms"].extend(normalized["chunk_terms"])

    merged_documents: list[dict] = []
    total_chunks = 0
    total_terms = 0

    for doc_id in sorted(documents.keys()):
        doc = documents[doc_id]
        chunk_seen: set[str] = set()
        chunk_terms: list[dict] = []

        for entry in doc.get("chunk_terms", []):
            chunk_id = entry["chunk_id"]
            if chunk_id in chunk_seen:
                continue
            chunk_seen.add(chunk_id)
            chunk_terms.append(entry)

        page_profile = merge_page_profiles(doc.get("page_profile", {}), (page_profile_index or {}).get(doc_id, {}))
        if page_profile:
            page_profile.setdefault("doc_id", doc_id)
        total_chunks += len(chunk_terms)
        total_terms += sum(len(entry["final_terms"]) for entry in chunk_terms)
        merged_documents.append(
            {
                "doc_id": doc_id,
                "page_profile": page_profile,
                "chunk_terms": chunk_terms,
            }
        )

    return {
        "merged_at": utc_now_iso(),
        "documents": merged_documents,
        "stats": {
            "document_count": len(merged_documents),
            "chunk_count": total_chunks,
            "term_count": total_terms,
        },
    }


def merge_llm_semantic_results(records: list[dict], page_profile_index: dict[str, dict] | None = None) -> dict:
    documents: dict[str, dict] = {}

    for record in records:
        normalized = normalize_semantic_doc_record(record, page_profile_index=page_profile_index)
        if not normalized:
            continue

        doc_id = normalized["doc_id"]
        slot = documents.setdefault(
            doc_id,
            {
                "doc_id": doc_id,
                "page_profile": normalized.get("page_profile", {}),
                "semantic_chunks": [],
            },
        )
        slot["page_profile"] = merge_page_profiles(slot.get("page_profile", {}), normalized.get("page_profile", {}))
        slot["semantic_chunks"].extend(normalized["semantic_chunks"])

    merged_documents: list[dict] = []
    total_chunks = 0

    for doc_id in sorted(documents.keys()):
        doc = documents[doc_id]
        chunk_seen: set[str] = set()
        semantic_chunks: list[dict] = []
        for entry in doc.get("semantic_chunks", []):
            chunk_id = entry["chunk_id"]
            if chunk_id in chunk_seen:
                continue
            chunk_seen.add(chunk_id)
            semantic_chunks.append(entry)

        page_profile = merge_page_profiles(doc.get("page_profile", {}), (page_profile_index or {}).get(doc_id, {}))
        if page_profile:
            page_profile.setdefault("doc_id", doc_id)
        total_chunks += len(semantic_chunks)
        merged_documents.append(
            {
                "doc_id": doc_id,
                "page_profile": page_profile,
                "semantic_chunks": semantic_chunks,
            }
        )

    return {
        "merged_at": utc_now_iso(),
        "documents": merged_documents,
        "stats": {
            "document_count": len(merged_documents),
            "chunk_count": total_chunks,
        },
    }


def write_semantic_excel_from_merged(merged_payload: dict, output_path: Path) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for --mode semantic-excel-export. "
            "Please install it first, e.g. `python3 -m pip install openpyxl`."
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "semantic_chunks"
    headers = [
        "doc_id",
        "source_file",
        "source_url",
        "marketplace",
        "page_title",
        "chunk_id",
        "heading",
        "heading_path",
        "screening_relevance_score",
        "screening_relevance_reason",
        "bucket_hints",
        "chunk_text",
        "chunk_source_url",
        "review_status",
        "review_notes",
        "review_chunk_patch",
    ]
    sheet.append(headers)

    for doc in merged_payload.get("documents", []):
        page_profile = doc.get("page_profile", {}) if isinstance(doc.get("page_profile"), dict) else {}
        for chunk in doc.get("semantic_chunks", []):
            sheet.append(
                [
                    doc.get("doc_id", ""),
                    page_profile.get("source_file", ""),
                    page_profile.get("source_url", ""),
                    page_profile.get("marketplace", ""),
                    page_profile.get("page_title", ""),
                    chunk.get("chunk_id", ""),
                    chunk.get("heading", ""),
                    chunk.get("heading_path", ""),
                    chunk.get("screening_relevance_score", ""),
                    chunk.get("screening_relevance_reason", ""),
                    " | ".join(unique_preserving_order([str(value) for value in chunk.get("bucket_hints", [])])),
                    chunk.get("chunk_text", ""),
                    chunk.get("chunk_source_url", ""),
                    "",
                    "",
                    "",
                ]
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_index_packets_from_semantic_documents(
    documents: list[dict],
    subagent_model: str | None = None,
    fallback_subagent_models: Any = None,
    min_index_score: int = DEFAULT_MIN_INDEX_SCORE,
) -> list[dict]:
    resolved_subagent_model = normalize_subagent_model(subagent_model)
    resolved_fallback_models = normalize_subagent_model_list(
        fallback_subagent_models,
        primary_model=resolved_subagent_model,
    )
    resolved_min_index_score = normalize_min_index_score(min_index_score)
    packets: list[dict] = []
    for doc in documents:
        doc_id = normalize_text(str(doc.get("doc_id", "")))
        if not doc_id:
            continue

        page_profile = doc.get("page_profile", {}) if isinstance(doc.get("page_profile"), dict) else {}
        page_entities_raw = page_profile.get("global_entities")
        page_entities = (
            unique_preserving_order([str(term) for term in page_entities_raw])
            if isinstance(page_entities_raw, list)
            else []
        )
        for chunk in doc.get("semantic_chunks", []):
            normalized_chunk = normalize_semantic_chunk_record(chunk)
            if not normalized_chunk:
                continue
            score = normalized_chunk.get("screening_relevance_score", 0)
            try:
                score = int(score)
            except (TypeError, ValueError):
                score = 0
            normalized_chunk["screening_relevance_score"] = score
            if score < resolved_min_index_score:
                continue
            if not normalized_chunk.get("bucket_hints"):
                normalized_chunk["bucket_hints"] = infer_regulation_bucket_hints(
                    normalized_chunk.get("heading", ""),
                    normalized_chunk.get("heading_path", ""),
                    normalized_chunk.get("chunk_text", ""),
                )

            heading = normalized_chunk.get("heading") or page_profile.get("page_title", "")
            seed_terms = build_seed_terms_for_chunk(
                chunk_text=normalized_chunk.get("chunk_text", ""),
                heading=heading,
                page_entities=page_entities,
            )
            packets.append(
                {
                    "doc_id": doc_id,
                    "chunk_id": normalized_chunk["chunk_id"],
                    "page_profile": page_profile,
                    "chunk_packet": normalized_chunk,
                    "seed_terms": {
                        "chunk_id": normalized_chunk["chunk_id"],
                        "seed_lang": "source+en",
                        **seed_terms,
                    },
                    "orchestration": {
                        "stage": "B",
                        "subagent_scope": "chunk",
                        "prompt_template_path": stage_b_prompt_template_path(),
                        "subagent_model": resolved_subagent_model,
                        "fallback_subagent_models": resolved_fallback_models,
                        "min_index_score": resolved_min_index_score,
                    },
                    "contract": {
                        "required_input": ["page_profile", "chunk_packet", "seed_terms"],
                        "required_output": {
                            "doc_id": "string",
                            "chunk_id": "string",
                            "final_terms": ["string"],
                        },
                        "notes": [
                            "逐 chunk 术语抽取；只输出 final_terms。",
                            "final_terms 需包含源语言术语与英文检索术语。",
                        ],
                    },
                }
            )
    return packets


def write_excel_from_merged(merged_payload: dict, output_path: Path) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for --mode excel-export. "
            "Please install it first, e.g. `python3 -m pip install openpyxl`."
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "chunk_terms"

    headers = [
        "doc_id",
        "source_file",
        "source_url",
        "marketplace",
        "page_title",
        "chunk_id",
        "final_terms",
        "final_terms_count",
        "review_status",
        "review_notes",
        "review_terms_patch",
        "reviewed_by",
        "reviewed_at",
    ]
    sheet.append(headers)

    documents = merged_payload.get("documents", [])
    for doc in documents:
        page_profile = doc.get("page_profile", {}) if isinstance(doc.get("page_profile"), dict) else {}
        for chunk in doc.get("chunk_terms", []):
            final_terms = normalize_term_list(chunk.get("final_terms"))
            sheet.append(
                [
                    doc.get("doc_id", ""),
                    page_profile.get("source_file", ""),
                    page_profile.get("source_url", ""),
                    page_profile.get("marketplace", ""),
                    page_profile.get("page_title", ""),
                    chunk.get("chunk_id", ""),
                    " | ".join(final_terms),
                    len(final_terms),
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def write_combined_review_excel(semantic_payload: dict, terms_payload: dict, output_path: Path) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required for --mode combined-review-export. "
            "Please install it first, e.g. `python3 -m pip install openpyxl`."
        ) from exc

    term_map: dict[tuple[str, str], list[str]] = {}
    for doc in terms_payload.get("documents", []):
        if not isinstance(doc, dict):
            continue
        doc_id = normalize_text(str(doc.get("doc_id", "")))
        for chunk in doc.get("chunk_terms", []):
            if not isinstance(chunk, dict):
                continue
            chunk_id = normalize_text(str(chunk.get("chunk_id", "")))
            if not doc_id or not chunk_id:
                continue
            term_map[(doc_id, chunk_id)] = normalize_term_list(chunk.get("final_terms"))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "combined_review"
    headers = [
        "doc_id",
        "page_title",
        "marketplace",
        "source_file",
        "source_url",
        "chunk_id",
        "heading",
        "heading_path",
        "screening_relevance_score",
        "screening_relevance_reason",
        "chunk_text",
        "final_terms",
        "final_terms_count",
        "review_status",
        "review_notes",
        "review_terms_patch",
    ]
    sheet.append(headers)

    for doc in semantic_payload.get("documents", []):
        if not isinstance(doc, dict):
            continue
        page_profile = doc.get("page_profile", {}) if isinstance(doc.get("page_profile"), dict) else {}
        doc_id = normalize_text(str(doc.get("doc_id", "")))
        for chunk in doc.get("semantic_chunks", []):
            if not isinstance(chunk, dict):
                continue
            chunk_id = normalize_text(str(chunk.get("chunk_id", "")))
            final_terms = term_map.get((doc_id, chunk_id), [])
            sheet.append(
                [
                    doc_id,
                    page_profile.get("page_title", ""),
                    page_profile.get("marketplace", ""),
                    page_profile.get("source_file", ""),
                    page_profile.get("source_url", ""),
                    chunk_id,
                    chunk.get("heading", ""),
                    chunk.get("heading_path", ""),
                    chunk.get("screening_relevance_score", ""),
                    chunk.get("screening_relevance_reason", ""),
                    chunk.get("chunk_text", ""),
                    " | ".join(final_terms),
                    len(final_terms),
                    "",
                    "",
                    "",
                ]
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def resolve_input_format(path: Path, input_format: str) -> str:
    if input_format != "auto":
        return input_format
    if path.suffix.lower() == ".md":
        return "markdown"
    return "html"


def iter_input_files(input_path: Path, input_format: str = "auto") -> Iterable[Path]:
    if input_path.is_file():
        yield input_path
        return

    if input_format == "html":
        patterns = ("*.html",)
    elif input_format == "markdown":
        patterns = ("*.md",)
    else:
        patterns = ("*.html", "*.md")

    found: list[Path] = []
    for pattern in patterns:
        found.extend(path for path in input_path.rglob(pattern) if path.is_file())
    yield from sorted(set(found))


def write_output(page: dict, output_dir: Path, mode: str = "draft") -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = Path(page["source_file"]).stem
    if mode == "chunks":
        output_path = output_dir / f"{stem}.chunks.md"
        content = render_semantic_chunks_markdown(page)
    else:
        output_path = output_dir / f"{stem}.md"
        content = render_extracted_markdown(page)
    output_path.write_text(content, encoding="utf-8")
    return output_path


def process_file(path: Path, input_format: str = "auto") -> dict:
    actual_format = resolve_input_format(path, input_format)
    if actual_format == "markdown":
        return parse_markdown_draft(path)

    html = read_html(path)
    root = extract_content_root(html)
    return extract_page_structure(root, path.name, source_url=resolve_source_url_for_file(path))


def build_arg_parser() -> argparse.ArgumentParser:
    default_output = default_output_root()
    parser = argparse.ArgumentParser(
        description="Extract Amazon help/regulation HTML or Markdown drafts into normalized outputs.",
    )
    parser.add_argument("--input", required=True, help="Input file or directory (HTML or Markdown draft).")
    parser.add_argument(
        "--input-format",
        choices=("auto", "html", "markdown"),
        default="auto",
        help="Input format detection mode. Default: auto.",
    )
    parser.add_argument(
        "--output-dir",
        help=f"Output directory. Default: {default_output}",
    )
    parser.add_argument(
        "--mode",
        choices=(
            "draft",
            "chunks",
            "index-seed",
            "llm-doc-packets",
            "llm-task-manifest",
            "llm-prompt-preview",
            "llm-runner-inputs",
            "llm-semantic-merge",
            "semantic-excel-export",
            "llm-index-packets",
            "llm-results-merge",
            "excel-export",
            "combined-review-export",
        ),
        default="draft",
        help="Output mode. Default: draft.",
    )
    parser.add_argument(
        "--index-output",
        help="JSON file path for --mode index-seed output.",
    )
    parser.add_argument(
        "--llm-packets-output",
        help="JSONL file path for --mode llm-doc-packets output.",
    )
    parser.add_argument(
        "--merged-output",
        help="JSON file path for merge modes output.",
    )
    parser.add_argument(
        "--excel-output",
        help="XLSX file path for excel export modes output.",
    )
    parser.add_argument(
        "--terms-input",
        help="JSON/JSONL path for Stage B term results when using --mode combined-review-export.",
    )
    parser.add_argument(
        "--index-packets-output",
        help="JSONL file path for --mode llm-index-packets output.",
    )
    parser.add_argument(
        "--manifest-output",
        help="JSON file path for --mode llm-task-manifest output.",
    )
    parser.add_argument(
        "--prompt-preview-dir",
        help="Directory path for --mode llm-prompt-preview output.",
    )
    parser.add_argument(
        "--runner-input-dir",
        help="Directory path for --mode llm-runner-inputs output.",
    )
    parser.add_argument(
        "--subagent-model",
        default=default_subagent_model(),
        help=(
            "Subagent model for Stage A/B orchestration metadata. "
            "Supported: GPT-5.3-Codex-Spark (default), GPT-5.4-Mini, GPT-5.4, GPT-5.3-Codex, GPT-5.2."
        ),
    )
    parser.add_argument(
        "--fallback-subagent-models",
        help=(
            "Comma-separated fallback models for capacity retry. "
            "Default: GPT-5.4-Mini when --subagent-model is GPT-5.3-Codex-Spark; otherwise none. "
            "Use 'none' to disable fallbacks."
        ),
    )
    parser.add_argument(
        "--min-index-score",
        type=int,
        default=DEFAULT_MIN_INDEX_SCORE,
        help=(
            "Minimum screening_relevance_score required for Stage B index packet generation. "
            f"Defaults to {DEFAULT_MIN_INDEX_SCORE}; chunks below this score are skipped."
        ),
    )
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    input_path = Path(args.input)
    input_format = args.input_format
    mode = args.mode
    output_root = Path(args.output_dir) if args.output_dir else default_output_root()
    index_output = Path(args.index_output) if args.index_output else None
    llm_packets_output = Path(args.llm_packets_output) if args.llm_packets_output else None
    merged_output = Path(args.merged_output) if args.merged_output else None
    excel_output = Path(args.excel_output) if args.excel_output else None
    terms_input = Path(args.terms_input) if args.terms_input else None
    index_packets_output = Path(args.index_packets_output) if args.index_packets_output else None
    manifest_output = Path(args.manifest_output) if args.manifest_output else None
    prompt_preview_dir = Path(args.prompt_preview_dir) if args.prompt_preview_dir else None
    runner_input_dir = Path(args.runner_input_dir) if args.runner_input_dir else None
    subagent_model = normalize_subagent_model(args.subagent_model)
    fallback_subagent_models = normalize_subagent_model_list(
        args.fallback_subagent_models,
        primary_model=subagent_model,
    )
    min_index_score = normalize_min_index_score(args.min_index_score)

    if not input_path.exists():
        raise SystemExit(f"Input path does not exist: {input_path}")

    if mode == "index-seed" and index_output is None:
        index_output = default_index_output_path(input_path, output_root)
    if mode == "llm-doc-packets" and llm_packets_output is None:
        llm_packets_output = default_llm_packets_output_path(input_path, output_root)
    if mode == "llm-task-manifest" and manifest_output is None:
        manifest_output = default_manifest_output_path(input_path, output_root)
    if mode == "llm-prompt-preview" and prompt_preview_dir is None:
        prompt_preview_dir = default_prompt_preview_dir(input_path, output_root)
    if mode == "llm-runner-inputs" and runner_input_dir is None:
        runner_input_dir = default_runner_input_dir(input_path, output_root)
    if mode == "llm-semantic-merge" and merged_output is None:
        merged_output = default_semantic_merged_output_path(input_path, output_root)
    if mode == "semantic-excel-export" and excel_output is None:
        excel_output = default_semantic_excel_output_path(input_path, output_root)
    if mode == "llm-index-packets" and index_packets_output is None:
        index_packets_output = default_index_packets_output_path(input_path, output_root)
    if mode == "llm-results-merge" and merged_output is None:
        merged_output = default_llm_merged_output_path(input_path, output_root)
    if mode == "excel-export" and excel_output is None:
        excel_output = default_excel_output_path(input_path, output_root)
    if mode == "combined-review-export" and excel_output is None:
        excel_output = default_combined_review_output_path(input_path, output_root)

    if mode == "combined-review-export":
        if terms_input is None:
            raise SystemExit("Mode 'combined-review-export' requires --terms-input.")
        if not terms_input.exists():
            raise SystemExit(f"Terms input path does not exist: {terms_input}")

    if mode in {"llm-task-manifest", "llm-prompt-preview", "llm-runner-inputs", "llm-semantic-merge", "semantic-excel-export", "llm-index-packets", "llm-results-merge", "excel-export", "combined-review-export"}:
        if input_path.is_dir():
            raise SystemExit(f"Mode '{mode}' expects --input to be a .json/.jsonl file, got directory: {input_path}")

        records = load_json_or_jsonl(input_path)
        if not records:
            raise SystemExit(f"No JSON records found in: {input_path}")
        page_profile_index = build_page_profile_index(
            input_path,
            terms_input if terms_input else None,
            output_root,
        )

        if mode == "llm-task-manifest":
            manifest = build_task_manifest(records, input_path, output_root)
            manifest_output.parent.mkdir(parents=True, exist_ok=True)
            manifest_output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
            print(manifest_output)
            return 0

        if mode == "llm-prompt-preview":
            preview_dir = write_prompt_previews(records, input_path, prompt_preview_dir)
            print(preview_dir)
            return 0

        if mode == "llm-runner-inputs":
            if prompt_preview_dir is None:
                prompt_preview_dir = default_prompt_preview_dir(input_path, output_root)
            runner_dir = write_runner_inputs(records, input_path, prompt_preview_dir, runner_input_dir)
            print(runner_dir if runner_input_dir is None else runner_input_dir)
            return 0

        if mode == "llm-semantic-merge":
            merged = merge_llm_semantic_results(records, page_profile_index=page_profile_index)
            if merged.get("stats", {}).get("document_count", 0) <= 0:
                raise SystemExit("No valid semantic chunk results found for merge.")
            merged_output.parent.mkdir(parents=True, exist_ok=True)
            merged_output.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
            print(merged_output)
            return 0

        if mode == "semantic-excel-export":
            merged_semantic = merge_llm_semantic_results(records, page_profile_index=page_profile_index)
            write_semantic_excel_from_merged(merged_semantic, excel_output)
            print(excel_output)
            return 0

        if mode == "llm-index-packets":
            merged_semantic = merge_llm_semantic_results(records, page_profile_index=page_profile_index)
            documents = merged_semantic.get("documents", [])
            packets = build_index_packets_from_semantic_documents(
                documents,
                subagent_model=subagent_model,
                fallback_subagent_models=fallback_subagent_models,
                min_index_score=min_index_score,
            )
            if not packets:
                raise SystemExit(
                    f"No semantic chunks found to build index packets at min_index_score={min_index_score}."
                )
            write_jsonl_records(packets, index_packets_output)
            print(index_packets_output)
            return 0

        if mode == "llm-results-merge":
            merged = merge_llm_doc_results(records, page_profile_index=page_profile_index)
            if merged.get("stats", {}).get("document_count", 0) <= 0:
                raise SystemExit("No valid document results found for merge.")
            merged_output.parent.mkdir(parents=True, exist_ok=True)
            merged_output.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
            print(merged_output)
            return 0

        if mode == "combined-review-export":
            semantic_payload = merge_llm_semantic_results(records, page_profile_index=page_profile_index)
            if semantic_payload.get("stats", {}).get("document_count", 0) <= 0:
                raise SystemExit("No valid semantic chunk results found for combined review export.")
            term_records = load_json_or_jsonl(terms_input)
            if not term_records:
                raise SystemExit(f"No term JSON records found in: {terms_input}")
            terms_payload = merge_llm_doc_results(term_records, page_profile_index=page_profile_index)
            if terms_payload.get("stats", {}).get("document_count", 0) <= 0:
                raise SystemExit("No valid term results found for combined review export.")
            write_combined_review_excel(semantic_payload, terms_payload, excel_output)
            print(excel_output)
            return 0

        merged_payload = merge_llm_doc_results(records, page_profile_index=page_profile_index)
        write_excel_from_merged(merged_payload, excel_output)
        print(excel_output)
        return 0

    file_paths = list(iter_input_files(input_path, input_format=input_format))
    if not file_paths:
        raise SystemExit(f"No input files found for format '{input_format}' in: {input_path}")
    pages = [process_file(file_path, input_format=input_format) for file_path in file_paths]

    if mode == "index-seed":
        payload = {
            "input_path": str(input_path),
            "pages": [build_index_seed_for_page(page) for page in pages],
        }
        index_output.parent.mkdir(parents=True, exist_ok=True)
        index_output.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(index_output)
        return 0

    if mode == "llm-doc-packets":
        packets = [
            build_llm_doc_packet_for_page(
                page,
                subagent_model=subagent_model,
                fallback_subagent_models=fallback_subagent_models,
            )
            for page in pages
        ]
        write_jsonl_records(packets, llm_packets_output)
        print(llm_packets_output)
        return 0

    for page in pages:
        output_path = write_output(page, output_root, mode=mode)
        print(output_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
