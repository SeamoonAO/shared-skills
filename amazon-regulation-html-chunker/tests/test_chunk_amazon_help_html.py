import importlib.util
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - optional dependency in test environment
    load_workbook = None


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "chunk_amazon_help_html.py"
FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"
SAMPLE_HTML = """
<html>
  <body>
    <div id="help-content">
      <h1>Policy Title</h1>
      <h2>Restricted Items</h2>
      <p>These products include batteries and chemicals for children.</p>
      <ul>
        <li><a name="sec_1"></a>Lithium battery pack</li>
        <li>Chemical cleaner for home kitchen</li>
      </ul>
      <h2>Allowed Items</h2>
      <p>General household accessories.</p>
    </div>
  </body>
</html>
""".strip()


def load_chunker_module():
    spec = importlib.util.spec_from_file_location("chunk_amazon_help_html", SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load module from {SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def run_cli(*args: str) -> subprocess.CompletedProcess:
    cmd = [sys.executable, str(SCRIPT_PATH), *args]
    return subprocess.run(cmd, check=True, capture_output=True, text=True)


class AmazonRegulationChunkerCliTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.module = load_chunker_module()

    def test_default_mode_generates_draft_markdown(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            index_path = root / "us_index.txt"
            index_path.write_text(
                "us_TestPolicy_G200164370.html\thttps://sellercentral.amazon.com/help/hub/reference/G200164370\n",
                encoding="utf-8",
            )

            output_dir = root / "out"
            run_cli("--input", str(html_path), "--output-dir", str(output_dir))

            output_path = output_dir / "us_TestPolicy_G200164370.md"
            self.assertTrue(output_path.exists())
            text = output_path.read_text(encoding="utf-8")
            self.assertIn("source_url: https://sellercentral.amazon.com/help/hub/reference/G200164370", text)
            self.assertIn("# Policy Title", text)

    def test_chunks_mode_generates_chunk_markdown(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            output_dir = root / "out"
            run_cli(
                "--input",
                str(html_path),
                "--output-dir",
                str(output_dir),
                "--mode",
                "chunks",
            )

            output_path = output_dir / "us_TestPolicy_G200164370.chunks.md"
            self.assertTrue(output_path.exists())
            text = output_path.read_text(encoding="utf-8")
            self.assertIn("## Chunk 1", text)
            self.assertIn("source_file: us_TestPolicy_G200164370.html", text)

    def test_index_seed_mode_single_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            index_output = root / "seed.json"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "index-seed",
                "--index-output",
                str(index_output),
            )

            self.assertTrue(index_output.exists())
            payload = json.loads(index_output.read_text(encoding="utf-8"))
            self.assertIn("pages", payload)
            self.assertEqual(len(payload["pages"]), 1)

            page_seed = payload["pages"][0]
            self.assertIn("page_profile", page_seed)
            self.assertIn("chunk_packet", page_seed)
            self.assertIn("seed_terms", page_seed)
            self.assertTrue(page_seed["chunk_packet"])

            first_chunk = page_seed["chunk_packet"][0]
            self.assertIn("chunk_id", first_chunk)
            self.assertIn("prev_context", first_chunk)
            self.assertIn("next_context", first_chunk)

            first_seed = page_seed["seed_terms"][0]
            self.assertNotIn("final_title", str(first_seed).lower())
            self.assertIn("category_terms", first_seed)

    def test_index_seed_mode_directory_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_dir = root / "html"
            html_dir.mkdir(parents=True, exist_ok=True)

            (html_dir / "us_TestPolicy_G200164370.html").write_text(SAMPLE_HTML, encoding="utf-8")
            (html_dir / "de_TestPolicy_G200164371.html").write_text(SAMPLE_HTML, encoding="utf-8")

            index_output = root / "batch_seed.json"
            run_cli(
                "--input",
                str(html_dir),
                "--mode",
                "index-seed",
                "--index-output",
                str(index_output),
            )

            payload = json.loads(index_output.read_text(encoding="utf-8"))
            self.assertEqual(len(payload["pages"]), 2)
            for page_seed in payload["pages"]:
                self.assertIn("page_profile", page_seed)
                self.assertIn("chunk_packet", page_seed)
                self.assertIn("seed_terms", page_seed)

    def test_default_output_dir_used_when_unspecified(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_DefaultOutputPolicy_G200199999.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            expected_output = self.module.default_output_root() / "us_DefaultOutputPolicy_G200199999.md"
            if expected_output.exists():
                expected_output.unlink()

            run_cli("--input", str(html_path))
            self.assertTrue(expected_output.exists())
            expected_output.unlink()

    def test_default_index_output_used_when_unspecified(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_dir = root / "html"
            html_dir.mkdir(parents=True, exist_ok=True)
            (html_dir / "us_DefaultSeedPolicy_G200188888.html").write_text(SAMPLE_HTML, encoding="utf-8")

            expected_index = self.module.default_index_output_path(
                html_dir,
                self.module.default_output_root(),
            )
            if expected_index.exists():
                expected_index.unlink()

            run_cli("--input", str(html_dir), "--mode", "index-seed")
            self.assertTrue(expected_index.exists())
            payload = json.loads(expected_index.read_text(encoding="utf-8"))
            self.assertIn("pages", payload)
            expected_index.unlink()

    def test_llm_doc_packets_mode_single_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            packets_output = root / "packets.jsonl"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
                "--subagent-model",
                "GPT-5.4-Mini",
            )

            self.assertTrue(packets_output.exists())
            lines = [line for line in packets_output.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(lines), 1)
            packet = json.loads(lines[0])
            self.assertIn("page_profile", packet)
            self.assertIn("chunk_packet", packet)
            self.assertIn("seed_terms", packet)
            self.assertIn("orchestration", packet)
            self.assertTrue(packet["chunk_packet"])
            self.assertTrue(packet["seed_terms"])
            self.assertEqual(packet["orchestration"]["subagent_model"], "gpt-5.4-mini")
            self.assertEqual(packet["orchestration"]["fallback_subagent_models"], [])
            self.assertEqual(packet["orchestration"]["stage"], "A")
            self.assertTrue(packet["orchestration"]["prompt_template_path"].endswith("subagent-stage-a-semantic-chunk-prompt.md"))
            self.assertIn("bucket_hints", packet["page_profile"])
            self.assertIn("bucket_hints", packet["chunk_packet"][0])

    def test_llm_doc_packets_default_capacity_fallback_for_spark(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            packets_output = root / "packets.jsonl"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
            )

            packet = json.loads(packets_output.read_text(encoding="utf-8").splitlines()[0])
            self.assertEqual(packet["orchestration"]["subagent_model"], "gpt-5.3-codex-spark")
            self.assertEqual(packet["orchestration"]["fallback_subagent_models"], ["gpt-5.4-mini"])

    def test_llm_doc_packets_can_disable_capacity_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            packets_output = root / "packets.jsonl"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
                "--fallback-subagent-models",
                "none",
            )

            packet = json.loads(packets_output.read_text(encoding="utf-8").splitlines()[0])
            self.assertEqual(packet["orchestration"]["fallback_subagent_models"], [])

    def test_llm_results_merge_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            results_input = root / "results.jsonl"
            merged_output = root / "merged.json"
            sample_records = [
                {
                    "doc_id": "us_G200164370",
                    "page_profile": {"source_file": "us_TestPolicy_G200164370.html", "marketplace": "us"},
                    "chunk_terms": [
                        {"chunk_id": "us_G200164370_c001", "final_terms": ["电池安全", "battery safety"]},
                        {"chunk_id": "us_G200164370_c001", "final_terms": ["重复项", "duplicate"]},
                    ],
                },
                {
                    "doc_id": "de_G201744080",
                    "page_profile": {"source_file": "de_TestPolicy_G201744080.html", "marketplace": "de"},
                    "chunk_terms": [
                        {"chunk_id": "de_G201744080_c001", "final_terms": ["Waffen", "weapon policy"]},
                    ],
                },
            ]
            results_input.write_text("\n".join(json.dumps(item, ensure_ascii=False) for item in sample_records), encoding="utf-8")

            run_cli(
                "--input",
                str(results_input),
                "--mode",
                "llm-results-merge",
                "--merged-output",
                str(merged_output),
            )

            self.assertTrue(merged_output.exists())
            merged = json.loads(merged_output.read_text(encoding="utf-8"))
            self.assertIn("documents", merged)
            self.assertEqual(merged["stats"]["document_count"], 2)
            self.assertEqual(merged["stats"]["chunk_count"], 2)

    def test_llm_task_manifest_mode_for_stage_a_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            packets_output = root / "packets.jsonl"
            manifest_output = root / "manifest.json"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
            )
            run_cli(
                "--input",
                str(packets_output),
                "--mode",
                "llm-task-manifest",
                "--manifest-output",
                str(manifest_output),
            )

            manifest = json.loads(manifest_output.read_text(encoding="utf-8"))
            self.assertEqual(manifest["task_count"], 1)
            self.assertEqual(manifest["stage_counts"]["A"], 1)
            self.assertEqual(manifest["tasks"][0]["stage"], "A")
            self.assertEqual(manifest["tasks"][0]["subagent_scope"], "document")
            self.assertEqual(manifest["tasks"][0]["subagent_model"], "gpt-5.3-codex-spark")
            self.assertEqual(manifest["tasks"][0]["fallback_subagent_models"], ["gpt-5.4-mini"])
            self.assertTrue(manifest["tasks"][0]["prompt_template_path"].endswith("subagent-stage-a-semantic-chunk-prompt.md"))
            self.assertTrue(manifest["tasks"][0]["preview_markdown_path"].endswith(".md"))
            self.assertTrue(manifest["tasks"][0]["runner_input_path"].endswith(".runner.md"))
            self.assertIn("semantic_chunks", manifest["tasks"][0]["expected_output"])
            self.assertIn("screening_relevance_score", manifest["tasks"][0]["expected_output"]["semantic_chunks"][0])

    def test_llm_prompt_preview_mode_for_stage_a_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            packets_output = root / "packets.jsonl"
            preview_dir = root / "preview"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
            )
            run_cli(
                "--input",
                str(packets_output),
                "--mode",
                "llm-prompt-preview",
                "--prompt-preview-dir",
                str(preview_dir),
            )

            previews = sorted(preview_dir.glob("*.md"))
            self.assertEqual(len(previews), 1)
            text = previews[0].read_text(encoding="utf-8")
            self.assertIn("## Regulation Source Markdown", text)
            self.assertNotIn("chunk_source_url", text)
            self.assertNotIn("prev_context", text)
            self.assertIn("Policy Title", text)

    def test_llm_runner_inputs_mode_for_stage_a_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            html_path = root / "us_TestPolicy_G200164370.html"
            html_path.write_text(SAMPLE_HTML, encoding="utf-8")

            packets_output = root / "packets.jsonl"
            runner_dir = root / "runner"
            run_cli(
                "--input",
                str(html_path),
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
            )
            run_cli(
                "--input",
                str(packets_output),
                "--mode",
                "llm-runner-inputs",
                "--runner-input-dir",
                str(runner_dir),
            )

            runner_files = sorted(runner_dir.glob("*.runner.md"))
            self.assertEqual(len(runner_files), 1)
            text = runner_files[0].read_text(encoding="utf-8")
            self.assertIn("## System Prompt Template", text)
            self.assertIn("\n---\n", text)
            self.assertIn("## Input Context", text)
            self.assertIn("## Execution Packet", text)
            self.assertIn("## Regulation Source Markdown", text)
            self.assertIn("## Runner Hints", text)
            self.assertIn("fallback_subagent_models: gpt-5.4-mini", text)
            self.assertNotIn("\"page_profile\"", text)
            self.assertNotIn("\"chunk_packet\"", text)

    def test_llm_runner_inputs_mode_for_stage_b_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            packets_output = root / "index-packets.jsonl"
            runner_dir = root / "runner"
            sample_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "doc_id": "us_G200164370",
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                    "global_entities": ["battery", "children"],
                },
                "semantic_chunks": [
                    {
                        "chunk_id": "us_G200164370_c001",
                        "heading": "Restricted Items",
                        "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries and chemicals for children.",
                        "chunk_source_url": "https://example.com/policy#c1",
                        "prev_context": "",
                        "next_context": "",
                        "bucket_hints": ["scope_actor", "domain_entity"],
                        "screening_relevance_score": 7,
                        "screening_relevance_reason": "screening-relevant restriction context",
                    }
                ],
            }
            semantic_input.write_text(json.dumps(sample_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "llm-index-packets",
                "--index-packets-output",
                str(packets_output),
            )
            run_cli(
                "--input",
                str(packets_output),
                "--mode",
                "llm-runner-inputs",
                "--runner-input-dir",
                str(runner_dir),
            )

            runner_files = sorted(runner_dir.glob("*.runner.md"))
            self.assertEqual(len(runner_files), 1)
            text = runner_files[0].read_text(encoding="utf-8")
            self.assertIn("## System Prompt Template", text)
            self.assertIn("\n---\n", text)
            self.assertIn("## Input Context", text)
            self.assertIn("## Execution Packet", text)
            self.assertIn("## Semantic Chunk", text)
            self.assertIn("## Seed Terms", text)
            self.assertIn("## Runner Hints", text)
            self.assertIn("fallback_subagent_models: gpt-5.4-mini", text)
            self.assertIn("screening_relevance_score", text)
            self.assertIn("bucket_hints", text)
            self.assertNotIn("\"page_profile\"", text)
            self.assertNotIn("\"chunk_packet\"", text)

    def test_llm_task_manifest_mode_for_stage_b_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            packets_output = root / "index-packets.jsonl"
            manifest_output = root / "manifest.json"
            sample_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "doc_id": "us_G200164370",
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                    "global_entities": ["battery", "children"],
                },
                "semantic_chunks": [
                    {
                        "chunk_id": "us_G200164370_c001",
                        "heading": "Restricted Items",
                        "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries and chemicals for children.",
                        "chunk_source_url": "https://example.com/policy#c1",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 7,
                        "screening_relevance_reason": "screening-relevant restriction context",
                    }
                ],
            }
            semantic_input.write_text(json.dumps(sample_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "llm-index-packets",
                "--index-packets-output",
                str(packets_output),
            )
            run_cli(
                "--input",
                str(packets_output),
                "--mode",
                "llm-task-manifest",
                "--manifest-output",
                str(manifest_output),
            )

            manifest = json.loads(manifest_output.read_text(encoding="utf-8"))
            self.assertEqual(manifest["task_count"], 1)
            self.assertEqual(manifest["stage_counts"]["B"], 1)
            self.assertEqual(manifest["tasks"][0]["stage"], "B")
            self.assertEqual(manifest["tasks"][0]["subagent_scope"], "chunk")
            self.assertEqual(manifest["tasks"][0]["subagent_model"], "gpt-5.3-codex-spark")
            self.assertEqual(manifest["tasks"][0]["fallback_subagent_models"], ["gpt-5.4-mini"])
            self.assertTrue(manifest["tasks"][0]["prompt_template_path"].endswith("subagent-stage-b-inverted-index-prompt.md"))
            self.assertTrue(manifest["tasks"][0]["preview_markdown_path"].endswith(".md"))
            self.assertTrue(manifest["tasks"][0]["runner_input_path"].endswith(".runner.md"))
            self.assertIn("final_terms", manifest["tasks"][0]["expected_output"])

    def test_default_runner_and_preview_dirs_strip_packet_suffixes(self) -> None:
        output_root = Path("/tmp/output-root")
        doc_packets = Path("/tmp/us-full.stage-a.doc-packets.jsonl")
        index_packets = Path("/tmp/live-stage-b-calibration.index-packets.jsonl")

        self.assertEqual(
            str(self.module.default_prompt_preview_dir(doc_packets, output_root)),
            "/tmp/output-root/us-full.stage-a.prompt-preview",
        )
        self.assertEqual(
            str(self.module.default_runner_input_dir(doc_packets, output_root)),
            "/tmp/output-root/us-full.stage-a.runner-inputs",
        )
        self.assertEqual(
            str(self.module.default_prompt_preview_dir(index_packets, output_root)),
            "/tmp/output-root/live-stage-b-calibration.prompt-preview",
        )
        self.assertEqual(
            str(self.module.default_runner_input_dir(index_packets, output_root)),
            "/tmp/output-root/live-stage-b-calibration.runner-inputs",
        )

    def test_llm_semantic_merge_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            merged_output = root / "semantic-merged.json"
            sample_records = [
                {
                    "doc_id": "us_G200164370",
                    "page_profile": {"source_file": "us_TestPolicy_G200164370.html", "marketplace": "us"},
                    "semantic_chunks": [
                        {
                            "chunk_id": "us_G200164370_c001",
                            "heading": "Restricted Items",
                            "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries.",
                        "chunk_source_url": "https://example.com#c1",
                        "prev_context": "",
                        "next_context": "Children use scenario",
                        "screening_relevance_score": 8,
                        "screening_relevance_reason": "explicit restriction example",
                    }
                ],
            },
                {
                    "doc_id": "de_G201744080",
                    "page_profile": {"source_file": "de_TestPolicy_G201744080.html", "marketplace": "de"},
                    "semantic_chunks": [
                        {
                            "chunk_id": "de_G201744080_c001",
                            "heading": "Waffen",
                            "heading_path": "Waffen",
                        "chunk_text": "Weapon-related restrictions.",
                        "chunk_source_url": "https://example.com/de#c1",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 9,
                        "screening_relevance_reason": "weapon restriction",
                    }
                ],
            },
            ]
            semantic_input.write_text("\n".join(json.dumps(item, ensure_ascii=False) for item in sample_records), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "llm-semantic-merge",
                "--merged-output",
                str(merged_output),
            )

            merged = json.loads(merged_output.read_text(encoding="utf-8"))
            self.assertEqual(merged["stats"]["document_count"], 2)
            self.assertEqual(merged["stats"]["chunk_count"], 2)
            self.assertIn("semantic_chunks", merged["documents"][0])
            first_chunk = merged["documents"][0]["semantic_chunks"][0]
            self.assertIn("screening_relevance_score", first_chunk)
            self.assertIn("screening_relevance_reason", first_chunk)

    def test_llm_index_packets_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            packets_output = root / "index-packets.jsonl"
            sample_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "doc_id": "us_G200164370",
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                    "global_entities": ["battery", "children"],
                },
                "semantic_chunks": [
                    {
                        "chunk_id": "us_G200164370_c001",
                        "heading": "Restricted Items",
                        "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries and chemicals for children.",
                        "chunk_source_url": "https://example.com/policy#c1",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 7,
                        "screening_relevance_reason": "screening-relevant restriction context",
                    },
                    {
                        "chunk_id": "us_G200164370_c002",
                        "heading": "Shipping Labels",
                        "heading_path": "Policy Title > Shipping Labels",
                        "chunk_text": "Generic shipping label formatting guidance.",
                        "chunk_source_url": "https://example.com/policy#c2",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 4,
                        "screening_relevance_reason": "low-value operational detail",
                    }
                ],
            }
            semantic_input.write_text(json.dumps(sample_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "llm-index-packets",
                "--index-packets-output",
                str(packets_output),
                "--subagent-model",
                "GPT-5.4",
            )

            lines = [line for line in packets_output.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(lines), 2)
            packet = json.loads(lines[0])
            self.assertIn("page_profile", packet)
            self.assertIn("chunk_packet", packet)
            self.assertIn("seed_terms", packet)
            self.assertIn("contract", packet)
            self.assertIn("orchestration", packet)
            self.assertEqual(packet["orchestration"]["stage"], "B")
            self.assertTrue(packet["orchestration"]["prompt_template_path"].endswith("subagent-stage-b-inverted-index-prompt.md"))
            self.assertEqual(packet["orchestration"]["subagent_model"], "gpt-5.4")
            self.assertEqual(packet["orchestration"]["fallback_subagent_models"], [])
            self.assertIn("bucket_hints", packet["chunk_packet"])
            self.assertIn("screening_relevance_score", packet["chunk_packet"])
            self.assertIn("screening_relevance_reason", packet["chunk_packet"])
            self.assertEqual(packet["chunk_packet"]["chunk_id"], "us_G200164370_c001")
            chunk_ids = [json.loads(line)["chunk_packet"]["chunk_id"] for line in lines]
            self.assertEqual(chunk_ids, ["us_G200164370_c001", "us_G200164370_c002"])

    def test_llm_index_packets_mode_respects_min_index_score_override(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            packets_output = root / "index-packets.jsonl"
            sample_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "doc_id": "us_G200164370",
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                    "global_entities": ["battery", "children"],
                },
                "semantic_chunks": [
                    {
                        "chunk_id": "us_G200164370_c001",
                        "heading": "Restricted Items",
                        "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries and chemicals for children.",
                        "chunk_source_url": "https://example.com/policy#c1",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 7,
                        "screening_relevance_reason": "screening-relevant restriction context",
                    },
                    {
                        "chunk_id": "us_G200164370_c002",
                        "heading": "Shipping Labels",
                        "heading_path": "Policy Title > Shipping Labels",
                        "chunk_text": "Generic shipping label formatting guidance.",
                        "chunk_source_url": "https://example.com/policy#c2",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 4,
                        "screening_relevance_reason": "low-value operational detail",
                    }
                ],
            }
            semantic_input.write_text(json.dumps(sample_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "llm-index-packets",
                "--index-packets-output",
                str(packets_output),
                "--min-index-score",
                "4",
            )

            lines = [line for line in packets_output.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(lines), 2)
            chunk_ids = [json.loads(line)["chunk_packet"]["chunk_id"] for line in lines]
            self.assertEqual(chunk_ids, ["us_G200164370_c001", "us_G200164370_c002"])

    @unittest.skipIf(load_workbook is None, "openpyxl is not installed")
    def test_semantic_excel_export_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            excel_output = root / "semantic.xlsx"
            sample_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                },
                "semantic_chunks": [
                    {
                        "chunk_id": "us_G200164370_c001",
                        "heading": "Restricted Items",
                        "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries and chemicals for children.",
                        "chunk_source_url": "https://example.com/policy#c1",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 7,
                        "screening_relevance_reason": "screening-relevant restriction context",
                    }
                ],
            }
            semantic_input.write_text(json.dumps(sample_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "semantic-excel-export",
                "--excel-output",
                str(excel_output),
            )

            workbook = load_workbook(excel_output)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("heading", headers)
            self.assertIn("chunk_text", headers)
            self.assertIn("bucket_hints", headers)
            self.assertIn("screening_relevance_score", headers)
            self.assertIn("screening_relevance_reason", headers)
            self.assertEqual(sheet.max_row, 2)

    @unittest.skipIf(load_workbook is None, "openpyxl is not installed")
    def test_semantic_excel_export_backfills_source_fields_from_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic-results.jsonl"
            packets_input = root / "stage-a.doc-packets.jsonl"
            excel_output = root / "semantic.xlsx"

            packets_input.write_text(
                json.dumps(
                    {
                        "doc_id": "us_G200164370",
                        "page_profile": {
                            "doc_id": "us_G200164370",
                            "source_file": "us_TestPolicy_G200164370.html",
                            "source_url": "https://example.com/policy",
                            "marketplace": "us",
                            "page_title": "Policy Title",
                        },
                        "chunk_packet": [{"chunk_id": "us_G200164370_u0001", "chunk_text": "raw packet"}],
                        "seed_terms": [],
                        "orchestration": {"stage": "A"},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            semantic_input.write_text(
                json.dumps(
                    {
                        "doc_id": "us_G200164370",
                        "semantic_chunks": [
                            {
                                "chunk_id": "us_G200164370_c001",
                                "heading": "Restricted Items",
                                "heading_path": "Policy Title > Restricted Items",
                                "chunk_text": "These products include batteries and chemicals for children.",
                                "chunk_source_url": "https://example.com/policy#c1",
                                "prev_context": "",
                                "next_context": "",
                                "screening_relevance_score": 7,
                                "screening_relevance_reason": "screening-relevant restriction context",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            run_cli(
                "--input",
                str(semantic_input),
                "--mode",
                "semantic-excel-export",
                "--excel-output",
                str(excel_output),
                "--output-dir",
                str(root),
            )

            workbook = load_workbook(excel_output)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            row = {headers[index]: sheet[2][index].value for index in range(len(headers))}
            self.assertEqual(row["source_file"], "us_TestPolicy_G200164370.html")
            self.assertEqual(row["source_url"], "https://example.com/policy")

    @unittest.skipIf(load_workbook is None, "openpyxl is not installed")
    def test_excel_export_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            results_input = root / "results.jsonl"
            excel_output = root / "result.xlsx"
            sample_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                },
                "chunk_terms": [
                    {"chunk_id": "us_G200164370_c001", "final_terms": ["电池安全", "battery safety"]},
                ],
            }
            results_input.write_text(json.dumps(sample_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(results_input),
                "--mode",
                "excel-export",
                "--excel-output",
                str(excel_output),
            )

            self.assertTrue(excel_output.exists())
            workbook = load_workbook(excel_output)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("chunk_id", headers)
            self.assertIn("final_terms", headers)
            self.assertEqual(sheet.max_row, 2)

    @unittest.skipIf(load_workbook is None, "openpyxl is not installed")
    def test_combined_review_export_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            terms_input = root / "terms.jsonl"
            excel_output = root / "combined-review.xlsx"

            semantic_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "doc_id": "us_G200164370",
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                },
                "semantic_chunks": [
                    {
                        "chunk_id": "us_G200164370_c001",
                        "heading": "Restricted Items",
                        "heading_path": "Policy Title > Restricted Items",
                        "chunk_text": "These products include batteries and chemicals for children.",
                        "chunk_source_url": "https://example.com/policy#c1",
                        "prev_context": "",
                        "next_context": "",
                        "screening_relevance_score": 7,
                        "screening_relevance_reason": "screening-relevant restriction context",
                    }
                ],
            }
            terms_record = {
                "doc_id": "us_G200164370",
                "page_profile": {
                    "doc_id": "us_G200164370",
                    "source_file": "us_TestPolicy_G200164370.html",
                    "source_url": "https://example.com/policy",
                    "marketplace": "us",
                    "page_title": "Policy Title",
                },
                "chunk_terms": [
                    {"chunk_id": "us_G200164370_c001", "final_terms": ["battery safety", "chemical cleaner"]},
                ],
            }
            semantic_input.write_text(json.dumps(semantic_record, ensure_ascii=False), encoding="utf-8")
            terms_input.write_text(json.dumps(terms_record, ensure_ascii=False), encoding="utf-8")

            run_cli(
                "--input",
                str(semantic_input),
                "--terms-input",
                str(terms_input),
                "--mode",
                "combined-review-export",
                "--excel-output",
                str(excel_output),
            )

            workbook = load_workbook(excel_output)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("source_url", headers)
            self.assertIn("screening_relevance_score", headers)
            self.assertIn("screening_relevance_reason", headers)
            self.assertIn("chunk_text", headers)
            self.assertIn("final_terms", headers)
            self.assertIn("final_terms_count", headers)
            self.assertEqual(sheet.max_row, 2)

    @unittest.skipIf(load_workbook is None, "openpyxl is not installed")
    def test_combined_review_export_backfills_source_fields_from_packets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            semantic_input = root / "semantic.jsonl"
            terms_input = root / "terms.jsonl"
            packets_input = root / "stage-a.doc-packets.jsonl"
            excel_output = root / "combined-review.xlsx"

            packets_input.write_text(
                json.dumps(
                    {
                        "doc_id": "us_G200164370",
                        "page_profile": {
                            "doc_id": "us_G200164370",
                            "source_file": "us_TestPolicy_G200164370.html",
                            "source_url": "https://example.com/policy",
                            "marketplace": "us",
                            "page_title": "Policy Title",
                        },
                        "chunk_packet": [{"chunk_id": "us_G200164370_u0001", "chunk_text": "raw packet"}],
                        "seed_terms": [],
                        "orchestration": {"stage": "A"},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            semantic_input.write_text(
                json.dumps(
                    {
                        "doc_id": "us_G200164370",
                        "semantic_chunks": [
                            {
                                "chunk_id": "us_G200164370_c001",
                                "heading": "Restricted Items",
                                "heading_path": "Policy Title > Restricted Items",
                                "chunk_text": "These products include batteries and chemicals for children.",
                                "chunk_source_url": "https://example.com/policy#c1",
                                "prev_context": "",
                                "next_context": "",
                                "screening_relevance_score": 7,
                                "screening_relevance_reason": "screening-relevant restriction context",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            terms_input.write_text(
                json.dumps(
                    {
                        "doc_id": "us_G200164370",
                        "chunk_terms": [
                            {"chunk_id": "us_G200164370_c001", "final_terms": ["battery safety", "chemical cleaner"]}
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            run_cli(
                "--input",
                str(semantic_input),
                "--terms-input",
                str(terms_input),
                "--mode",
                "combined-review-export",
                "--excel-output",
                str(excel_output),
                "--output-dir",
                str(root),
            )

            workbook = load_workbook(excel_output)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            row = {headers[index]: sheet[2][index].value for index in range(len(headers))}
            self.assertEqual(row["source_file"], "us_TestPolicy_G200164370.html")
            self.assertEqual(row["source_url"], "https://example.com/policy")


class AmazonRegulationChunkerFixtureTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.us_fixtures = sorted((FIXTURES_DIR / "amazon-drafts-us").glob("*.md"))
        cls.de_fixtures = sorted((FIXTURES_DIR / "amazon-drafts-de").glob("*.md"))
        if not cls.us_fixtures or not cls.de_fixtures:
            raise RuntimeError("Fixture drafts not found for US/DE.")
        cls.us_primary = FIXTURES_DIR / "amazon-drafts-us" / "us_动物和动物相关商品_G200164370.md"
        cls.us_secondary = FIXTURES_DIR / "amazon-drafts-us" / "us_食品和饮料_G200164550.md"
        cls.de_primary = FIXTURES_DIR / "amazon-drafts-de" / "de_武器_G201744080.md"
        cls.de_secondary = FIXTURES_DIR / "amazon-drafts-de" / "de_动物和动物制品_G201743950.md"
        required = [cls.us_primary, cls.us_secondary, cls.de_primary, cls.de_secondary]
        missing = [str(path) for path in required if not path.exists()]
        if missing:
            raise RuntimeError(f"Required fixture files missing: {missing}")

    def test_fixture_markdown_input_index_seed_single_file(self) -> None:
        fixture_path = self.us_primary
        with tempfile.TemporaryDirectory() as tmp:
            index_output = Path(tmp) / "seed.json"
            run_cli(
                "--input",
                str(fixture_path),
                "--input-format",
                "markdown",
                "--mode",
                "index-seed",
                "--index-output",
                str(index_output),
            )
            payload = json.loads(index_output.read_text(encoding="utf-8"))
            self.assertEqual(len(payload["pages"]), 1)

            page_seed = payload["pages"][0]
            self.assertIn("page_profile", page_seed)
            self.assertIn("chunk_packet", page_seed)
            self.assertIn("seed_terms", page_seed)
            self.assertTrue(page_seed["chunk_packet"])

            first_seed = page_seed["seed_terms"][0]
            self.assertNotIn("final_title", str(first_seed).lower())
            self.assertIn("category_terms", first_seed)
            self.assertIn("risk_terms", first_seed)
            self.assertIn("scenario_terms", first_seed)
            self.assertIn("variant_terms", first_seed)

    def test_fixture_markdown_input_chunks_mode_single_file(self) -> None:
        fixture_path = self.de_primary
        with tempfile.TemporaryDirectory() as tmp:
            output_dir = Path(tmp) / "out"
            run_cli(
                "--input",
                str(fixture_path),
                "--input-format",
                "markdown",
                "--mode",
                "chunks",
                "--output-dir",
                str(output_dir),
            )

            expected_stem = fixture_path.stem
            output_path = output_dir / f"{expected_stem}.chunks.md"
            self.assertTrue(output_path.exists())
            text = output_path.read_text(encoding="utf-8")
            self.assertIn("## Chunk 1", text)
            self.assertIn("source_file:", text)

    def test_fixture_markdown_input_index_seed_directory_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture_dir = root / "subset"
            fixture_dir.mkdir(parents=True, exist_ok=True)

            selected = [self.us_primary, self.us_secondary, self.de_primary, self.de_secondary]
            for src in selected:
                shutil.copy2(src, fixture_dir / src.name)

            index_output = root / "batch_seed.json"
            run_cli(
                "--input",
                str(fixture_dir),
                "--input-format",
                "markdown",
                "--mode",
                "index-seed",
                "--index-output",
                str(index_output),
            )

            payload = json.loads(index_output.read_text(encoding="utf-8"))
            self.assertEqual(len(payload["pages"]), len(selected))
            for page_seed in payload["pages"]:
                self.assertIn("page_profile", page_seed)
                self.assertIn("chunk_packet", page_seed)
                self.assertIn("seed_terms", page_seed)
                self.assertGreaterEqual(len(page_seed["chunk_packet"]), 1)

    def test_fixture_markdown_input_llm_doc_packets_directory_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture_dir = root / "subset"
            fixture_dir.mkdir(parents=True, exist_ok=True)

            selected = [self.us_primary, self.de_primary]
            for src in selected:
                shutil.copy2(src, fixture_dir / src.name)

            packets_output = root / "packets.jsonl"
            run_cli(
                "--input",
                str(fixture_dir),
                "--input-format",
                "markdown",
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(packets_output),
            )

            lines = [line for line in packets_output.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(lines), len(selected))
            for line in lines:
                payload = json.loads(line)
                self.assertIn("page_profile", payload)
                self.assertIn("chunk_packet", payload)
                self.assertIn("seed_terms", payload)
                self.assertTrue(payload["chunk_packet"])
                self.assertIn("orchestration", payload)

    def test_all_fixtures_markdown_llm_doc_packets_smoke(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "all-fixtures.jsonl"
            run_cli(
                "--input",
                str(FIXTURES_DIR),
                "--input-format",
                "markdown",
                "--mode",
                "llm-doc-packets",
                "--llm-packets-output",
                str(output_path),
            )

            expected_count = len(sorted(FIXTURES_DIR.rglob("*.md")))
            lines = [line for line in output_path.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(lines), expected_count)


if __name__ == "__main__":
    unittest.main()
